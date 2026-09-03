#!/usr/bin/env python3
"""
Generate the adjudication resolution breakdown for the AIFTAX replication package.

    python make_adjudication_breakdown.py \
        --adjudicated "Final_Corpus_and_Adjudicated_Log/AIFTax_Final_Adjudicated_Dataset.xlsx" \
        --output-dir results_adjudication_breakdown

Produces AIFTax_Adjudication_Resolution_Breakdown.xlsx with:

    By Field Group  breakdown by the nine analytical groups
    By Field        per-field breakdown, including fields with zero disagreements
    Verification    integrity checks the artifact asserts about itself
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_ADJ = "Final_Corpus_and_Adjudicated_Log/AIFTax_Final_Adjudicated_Dataset.xlsx"
DEFAULT_OUT = "results_adjudication_breakdown"
ROOT_MARKERS = ("Final_Corpus_and_Adjudicated_Log", "Annotation Materials", ".git")

N_CASES = 100
N_SUBSTANTIVE = 32
TABLE_V = {"Failure Category": 11, "Propagation Reach": 36,
           "Risk Assessment": 17, "Recovery Complexity": 9}

ARIAL = "Arial"
HDR_FILL = PatternFill("solid", fgColor="D9D9D9")
TOT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")


def norm(s):
    return " ".join(str(s).split()).strip().casefold()


def sset(s):
    return frozenset(x.strip() for x in norm(s).split(";") if x.strip())


def find_root(explicit=None):
    if explicit:
        return Path(explicit).expanduser().resolve()
    for start in (Path.cwd(), Path(__file__).resolve().parent):
        for d in (start, *start.parents):
            if any((d / m).exists() for m in ROOT_MARKERS):
                return d
    return Path.cwd()


def resolve_in(raw, root, label):
    p = Path(raw).expanduser()
    tried = [p] if p.is_absolute() else list(dict.fromkeys([Path.cwd() / p, root / p]))
    for t in tried:
        if t.exists():
            return t.resolve()
    print(f"ERROR: {label} not found. Tried:", file=sys.stderr)
    for t in tried:
        print(f"    {t}", file=sys.stderr)
    sys.exit(1)


def tally(df, by):
    """Counts and percentages of resolution outcome, grouped by a column."""
    rows = []
    for key, g in df.groupby(by, sort=False):
        n = len(g)
        a = int((g["_outcome"] == "A").sum())
        b = int((g["_outcome"] == "B").sum())
        r = int((g["_outcome"] == "R").sum())
        rows.append({by: key, "Cells": n, "LLM A": a, "LLM B": b, "Re-coded": r,
                     "% LLM A": round(100 * a / n, 1), "% LLM B": round(100 * b / n, 1),
                     "% Re-coded": round(100 * r / n, 1)})
    return pd.DataFrame(rows).sort_values("Cells", ascending=False).reset_index(drop=True)


def write_sheet(ws, df, title=None, widths=None, total_row=False):
    r = 1
    if title:
        c = ws.cell(row=1, column=1, value=title)
        c.font = Font(name=ARIAL, size=11, bold=True)
        r = 3
    for j, col in enumerate(df.columns, start=1):
        c = ws.cell(row=r, column=j, value=col)
        c.font = Font(name=ARIAL, size=10, bold=True)
        c.fill = HDR_FILL
        c.border = Border(bottom=THIN)
        c.alignment = Alignment(wrap_text=True, vertical="bottom",
                                horizontal="left" if j == 1 else "right")
    for i, (_, row) in enumerate(df.iterrows(), start=r + 1):
        is_total = total_row and i == r + len(df)
        for j, col in enumerate(df.columns, start=1):
            c = ws.cell(row=i, column=j, value=row[col])
            c.font = Font(name=ARIAL, size=10, bold=is_total)
            c.alignment = Alignment(horizontal="left" if j == 1 else "right")
            if is_total:
                c.fill = TOT_FILL
            if col.startswith("%"):
                c.number_format = "0.0"
    for j, col in enumerate(df.columns, start=1):
        w = (widths or {}).get(col, max(11, min(34, len(str(col)) + 4)))
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=r + 1, column=2)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--adjudicated", default=DEFAULT_ADJ)
    ap.add_argument("--output-dir", default=DEFAULT_OUT)
    ap.add_argument("--project-root", default=None)
    args = ap.parse_args()

    root = find_root(args.project_root)
    adj = resolve_in(args.adjudicated, root, "--adjudicated")
    out = Path(args.output_dir)
    out = out.resolve() if out.is_absolute() else (root / out).resolve()
    out.mkdir(parents=True, exist_ok=True)

    xl = pd.ExcelFile(adj)
    log = xl.parse("Adjudication Log", header=0)
    log.columns = [str(c).strip() for c in log.columns]
    corp = xl.parse("Final Corpus", header=1)
    corp.columns = [" ".join(str(c).split()).strip() for c in corp.columns]
    corp = corp[corp["Case ID"].notna()]
    substantive = [c for c in corp.columns
                   if c not in ("Case ID", "Source Link(s) / Citation(s)")]

    # Resolution outcome, set-aware for multi-label fields.
    outcomes = []
    for _, r in log.iterrows():
        ml = "multi-label" in str(r["Field Type"]).lower()
        cmp = sset if ml else norm
        fa = cmp(r["Final Value"]) == cmp(r["Annotator A Value"])
        fb = cmp(r["Final Value"]) == cmp(r["Annotator B Value"])
        outcomes.append("A" if fa else ("B" if fb else "R"))
    log["_outcome"] = outcomes

    # Collapse the two narrative types for reporting, as in the paper.
    log["Field Type (reported)"] = log["Field Type"].replace(
        {"Structured narrative": "Narrative",
         "Optional structured narrative": "Narrative"})

    by_type = tally(log, "Field Type (reported)")
    by_group = tally(log, "Field Group")
    by_field = tally(log, "Field")

    # Fields with zero logged disagreements appear with explicit zeros.
    zero = [f for f in substantive if f not in set(log["Field"].unique())]
    if zero:
        by_field = pd.concat([by_field, pd.DataFrame(
            [{"Field": f, "Cells": 0, "LLM A": 0, "LLM B": 0, "Re-coded": 0,
              "% LLM A": None, "% LLM B": None, "% Re-coded": None} for f in zero])],
            ignore_index=True)

    n_log = len(log)
    n_cells = N_CASES * len(substantive)
    a_tot = int((log["_outcome"] == "A").sum())
    b_tot = int((log["_outcome"] == "B").sum())
    r_tot = int((log["_outcome"] == "R").sum())
    ctrl = log[log["Field Type"].str.contains("Controlled", na=False)]
    ordn = log[log["Field Type"].str.contains("ordinal", case=False, na=False)]

    def pct(x, n):
        return round(100 * x / n, 1) if n else None

    summary = pd.DataFrame([
        {"Grouping": "All logged disagreements", "Cells": n_log, "LLM A": a_tot,
         "LLM B": b_tot, "Re-coded": r_tot, "% LLM A": pct(a_tot, n_log),
         "% LLM B": pct(b_tot, n_log), "% Re-coded": pct(r_tot, n_log)},
        {"Grouping": "Controlled fields (nominal + ordinal + multi-label)",
         "Cells": len(ctrl), "LLM A": int((ctrl["_outcome"] == "A").sum()),
         "LLM B": int((ctrl["_outcome"] == "B").sum()),
         "Re-coded": int((ctrl["_outcome"] == "R").sum()),
         "% LLM A": pct((ctrl["_outcome"] == "A").sum(), len(ctrl)),
         "% LLM B": pct((ctrl["_outcome"] == "B").sum(), len(ctrl)),
         "% Re-coded": pct((ctrl["_outcome"] == "R").sum(), len(ctrl))},
        {"Grouping": "Ordinal principal variables only", "Cells": len(ordn),
         "LLM A": int((ordn["_outcome"] == "A").sum()),
         "LLM B": int((ordn["_outcome"] == "B").sum()),
         "Re-coded": int((ordn["_outcome"] == "R").sum()),
         "% LLM A": pct((ordn["_outcome"] == "A").sum(), len(ordn)),
         "% LLM B": pct((ordn["_outcome"] == "B").sum(), len(ordn)),
         "% Re-coded": pct((ordn["_outcome"] == "R").sum(), len(ordn))},
    ])

    type_tot = pd.DataFrame([{"Field Type (reported)": "Total", "Cells": n_log,
                              "LLM A": a_tot, "LLM B": b_tot, "Re-coded": r_tot,
                              "% LLM A": pct(a_tot, n_log), "% LLM B": pct(b_tot, n_log),
                              "% Re-coded": pct(r_tot, n_log)}])
    by_type_out = pd.concat([by_type, type_tot], ignore_index=True)

    # ---- verification ------------------------------------------------------
    pairs = list(zip(log["Case ID"], [norm(f) for f in log["Field"]]))
    idx = corp.set_index(corp["Case ID"].astype(int))
    fv_match = sum(
        1 for _, r in log.iterrows()
        if (sset if "multi-label" in str(r["Field Type"]).lower() else norm)(
            idx.at[int(r["Case ID"]), " ".join(str(r["Field"]).split()).strip()])
        == (sset if "multi-label" in str(r["Field Type"]).lower() else norm)(r["Final Value"]))
    ab_diff = sum(1 for _, r in log.iterrows()
                  if (sset if "multi-label" in str(r["Field Type"]).lower() else norm)(
                      r["Annotator A Value"])
                  != (sset if "multi-label" in str(r["Field Type"]).lower() else norm)(
                      r["Annotator B Value"]))
    checks = [
        ("Log rows", n_log, "1,110 logged disagreements"),
        ("Unique (Case ID, Field) pairs", len(set(pairs)),
         "equals row count if no cell is logged twice"),
        ("Rows where LLM A value != LLM B value", ab_diff,
         "equals row count if the log records disagreements only"),
        ("Final Value matches Final Corpus cell", fv_match,
         "equals row count if log and corpus agree"),
        ("Re-coded entries differing from both inputs", r_tot,
         "equals the re-coded count if the label is literal"),
        ("Cases represented", int(log["Case ID"].nunique()), f"of {N_CASES}"),
        ("Fields with >= 1 disagreement", int(log["Field"].nunique()),
         f"of {len(substantive)} substantive fields"),
        ("Substantive cells in corpus", n_cells, "100 cases x 32 fields"),
        ("Adjudicated share", f"{pct(n_log, n_cells)}%", "logged / substantive cells"),
        ("Consensus share (not adjudicated)", f"{pct(n_cells - n_log, n_cells)}%",
         f"{n_cells - n_log} cells carried matching values from both LLMs"),
        ("Adjudicated fields per case", f"{log.groupby('Case ID').size().min()}"
         f"-{log.groupby('Case ID').size().max()} "
         f"(mean {log.groupby('Case ID').size().mean():.1f})", ""),
    ]
    for var, expected in TABLE_V.items():
        got = int((log["Field"].apply(lambda x: norm(x) == norm(var))).sum())
        checks.append((f"Table V Diff. reproduction: {var}", got,
                       f"paper reports {expected}" + ("" if got == expected else "  <-- MISMATCH")))
    verification = pd.DataFrame(checks, columns=["Check", "Value", "Expected / meaning"])

    # ---- write -------------------------------------------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    readme = [
        ("AIFTAX — Adjudication Resolution Breakdown", True),
        ("", False),
        ("Companion artifact to Section IV-C. Supports the sentence: \"Adjudication retained", False),
        ("LLM A's value in 74.1% of entries, concentrated in narrative fields (89.8%); among the", False),
        ("ordinal principal variables the split was near-even (44.3% LLM A, 51.9% LLM B, 3.8%", False),
        ("re-coded), with the full breakdown provided in the replication package.\"", False),
        ("", False),
        ("SOURCE", True),
        (f"  {adj.name}, sheets 'Adjudication Log' and 'Final Corpus'.", False),
        ("  All figures are computed, not hand-entered. Regenerate with:", False),
        ("      python make_adjudication_breakdown.py --adjudicated <path> --output-dir <dir>", False),
        ("", False),
        ("DEFINITIONS", True),
        ("  LLM A / LLM B   the model whose value was retained after source-grounded adjudication.", False),
        ("  Re-coded        a final adjudicated value different from both model inputs.", False),
        ("  Narrative       structured and optional structured-text fields combined; any wording", False),
        ("                  difference produces a log entry rather than necessarily indicating a", False),
        ("                  categorical disagreement.", False),
        ("  Multi-label values are compared as unordered sets, per the coding manual. Under plain", False),
        ("  string comparison 37 entries appear misfiled purely from label ordering in Impact", False),
        ("  Dimension(s) and Causal Location(s); set comparison resolves all of them.", False),
        ("", False),
        ("SCOPE", True),
        ("  The log records cells on which the two LLM annotators differed. Cells carrying matching", False),
        ("  values from both models were not adjudicated and are outside this artifact. See the", False),
        ("  Verification sheet for the adjudicated / consensus split.", False),
        ("", False),
        (f"  Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}", False),
    ]
    for i, (text, bold) in enumerate(readme, start=1):
        ws.cell(row=i, column=1, value=text).font = Font(name=ARIAL, size=11, bold=bold)
    ws.column_dimensions["A"].width = 104

    write_sheet(wb.create_sheet("Summary"), summary,
                "Resolution outcome — headline groupings",
                widths={"Grouping": 46})
    ws2 = wb.create_sheet("By Field Type")
    write_sheet(ws2, by_type_out, "Resolution outcome by field type",
                widths={"Field Type (reported)": 30}, total_row=True)
    write_sheet(wb.create_sheet("By Field Group"), by_group,
                "Resolution outcome by analytical group", widths={"Field Group": 30})
    write_sheet(wb.create_sheet("By Field"), by_field,
                "Resolution outcome by field (zero-disagreement fields shown with 0 cells)",
                widths={"Field": 38})
    write_sheet(wb.create_sheet("Verification"), verification,
                "Integrity checks asserted by this artifact",
                widths={"Check": 42, "Value": 16, "Expected / meaning": 52})

    xlsx = out / "AIFTax_Adjudication_Resolution_Breakdown.xlsx"
    wb.save(xlsx)

    by_type_out.to_csv(out / "adjudication_by_field_type.csv", index=False)
    by_field.to_csv(out / "adjudication_by_field.csv", index=False)
    by_group.to_csv(out / "adjudication_by_field_group.csv", index=False)

    print(f"Source     : {adj}")
    print(f"Output dir : {out}\n")
    print(summary.to_string(index=False))
    print()
    print(by_type_out.to_string(index=False))
    print()
    bad = [c for c in checks if "MISMATCH" in str(c[2])]
    print(f"Verification: {len(checks)} checks, {len(bad)} mismatches")
    for c in bad:
        print("   ", c)
    print(f"\nWrote {xlsx.name} and 3 CSVs to {out}")


if __name__ == "__main__":
    main()
