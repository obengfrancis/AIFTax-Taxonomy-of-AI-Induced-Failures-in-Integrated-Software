#!/usr/bin/env python3
"""
AIFTax held-out inter-model agreement analysis.

Compares two independent LLM annotators on the same out-of-sample incidents.

Recommended use:
    python aiftax_heldout_agreement.py \
      --annotator-c "Annotation Materials/AIFTax_Annotator_C_45_Cases_Completed.xlsx" \
      --annotator-d "Annotation Materials/AIFTax_Annotator_D_45_Cases_Completed.xlsx" \
      --output-dir results_heldout_agreement

The script reports TWO analysis populations when --exclude-case-ids is supplied:
IMPORTANT:
- --exclude-case-ids is an external screening input. The script NEVER derives
  eligibility from either annotator's Failure Category or other annotation.
- IE/NA are retained as legitimate annotation outcomes for unweighted agreement.
- For ordinal quadratic-weighted kappa, pairs containing IE/NA/blank are excluded
  because IE/NA have no ordinal position.
- Free-text fields are not assigned a content kappa. Their IE/NA/blank status is
  still included in the missingness-agreement diagnostics.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# Field schema
# ---------------------------------------------------------------------------

ADMIN_FIELDS = {"Case ID", "Source Link(s) / Citation(s)"}
NOTES_MARKER = "Notes"

MISSING_CODES = {
    "ie", "na", "n/a", "insufficient evidence", "not applicable"
}

# Formal reliability fields.
# These are controlled-vocabulary fields for which categorical agreement is
# meaningful. Open/structured narrative fields are handled only in the
# missingness diagnostics.
NOMINAL_FIELDS = [
    "System Type",
    "Industry",
    "Failure Category",
    "Manifestation Pattern",
    "Boundary Transfer",
    "Transfer Mode",
    "Amplification",
    "Detection Timing",
    "Detector",
    "Detection Signal",
    "Risk Domain",
    "Recovery Evidence Status",
    "Lifecycle Phase",
    "Temporal Pattern",
]

ORDINAL_FIELDS = {
    "Propagation Reach": ["P0", "P1", "P2", "P3"],
    "Risk Assessment": ["Low", "Moderate", "High", "Severe", "Critical"],
    "Recovery Complexity": ["Low", "Moderate", "High", "Severe", "Very High"],
}

# Failure Mode(s) routinely carries two or more labels per incident, so it is
# compared as a set like the other multi-label fields.
MULTILABEL_FIELDS = [
    "Failure Mode(s)",
    "Impact Dimension(s)",
    "Recovery / Maintenance Action(s)",
    "Causal Location(s)",
]

FORMAL_FIELDS = NOMINAL_FIELDS + list(ORDINAL_FIELDS) + MULTILABEL_FIELDS

PRINCIPAL_FIELDS = [
    "Failure Category",
    "Propagation Reach",
    "Risk Assessment",
    "Recovery Complexity",
]


# ---------------------------------------------------------------------------
# Workbook loading and normalization
# ---------------------------------------------------------------------------

def clean_header(value) -> str:
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def load_records(path: Path, sheet: str = "Annotations"):
    """Load the annotation sheet and return (headers, records)."""
    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{path.name}: missing sheet {sheet!r}")

    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_index = None
    headers = None
    for i, row in enumerate(rows[:25]):
        cells = [clean_header(v) for v in row]
        if "Case ID" in cells:
            header_index = i
            headers = cells
            break

    if header_index is None or headers is None:
        raise ValueError(f"{path.name}: no header row containing 'Case ID'")

    records = []
    for row in rows[header_index + 1:]:
        if not row:
            continue
        rec = dict(zip(headers, row))
        cid = rec.get("Case ID")
        if cid is None or not str(cid).strip():
            continue
        records.append(rec)

    return headers, records


def case_id(value) -> str:
    if value is None:
        return ""
    # Excel may return 5 or 5.0 depending on workbook construction.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def raw_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_value(field: str, value) -> str:
    """Normalize a single controlled value for agreement comparison."""
    text = raw_text(value)
    if not text:
        return ""

    # Normalize whitespace.
    text = re.sub(r"\s+", " ", text).strip()

    if field == "Propagation Reach":
        m = re.match(r"^(P[0-3])\b", text, re.I)
        if m:
            return m.group(1).upper()
        if text.casefold() in MISSING_CODES:
            return "IE" if text.casefold().startswith("ie") or "insufficient" in text.casefold() else "NA"

    # Normalize common missing-code spellings.
    cf = text.casefold()
    if cf in MISSING_CODES:
        if cf in {"na", "n/a", "not applicable"}:
            return "NA"
        return "IE"

    return text


def split_multilabel(field: str, value) -> frozenset[str]:
    text = raw_text(value)
    if not text:
        return frozenset()

    parts = []
    for part in text.split(";"):
        label = normalize_value(field, part)
        if label:
            parts.append(label)
    return frozenset(parts)


def is_missing_code(value) -> bool:
    text = raw_text(value)
    return bool(text) and text.casefold() in MISSING_CODES


def is_blank(value) -> bool:
    return not raw_text(value)


def evidence_status(value) -> str:
    """Three-state diagnostic status: coded, IE/NA, or blank."""
    if is_blank(value):
        return "blank"
    if is_missing_code(value):
        return "IE/NA"
    return "coded"


def substantive_fields(headers) -> list[str]:
    return [
        h for h in headers
        if h and h not in ADMIN_FIELDS and NOTES_MARKER not in h
    ]


# ---------------------------------------------------------------------------
# Agreement statistics
# ---------------------------------------------------------------------------

def cohen_kappa(pairs) -> float | None:
    """Unweighted Cohen's kappa for arbitrary hashable category pairs."""
    n = len(pairs)
    if n == 0:
        return None

    observed = sum(1 for a, b in pairs if a == b) / n
    ca = Counter(a for a, _ in pairs)
    cb = Counter(b for _, b in pairs)
    keys = set(ca) | set(cb)
    expected = sum((ca[k] / n) * (cb[k] / n) for k in keys)
    denominator = 1.0 - expected

    if math.isclose(denominator, 0.0, abs_tol=1e-12):
        # Perfect agreement with no marginal variance.
        return 1.0 if math.isclose(observed, 1.0, abs_tol=1e-12) else None

    return (observed - expected) / denominator


def quadratic_weighted_kappa(pairs, ordered_labels: list[str]) -> float | None:
    """
    Quadratic-weighted Cohen's kappa.

    Only pairs whose BOTH labels occur in ordered_labels should be passed here.
    """
    if not pairs:
        return None

    index = {label: i for i, label in enumerate(ordered_labels)}
    k = len(ordered_labels)
    if k < 2:
        return None

    n = len(pairs)
    obs = [[0.0] * k for _ in range(k)]
    a_counts = [0.0] * k
    b_counts = [0.0] * k

    for a, b in pairs:
        i, j = index[a], index[b]
        obs[i][j] += 1.0
        a_counts[i] += 1.0
        b_counts[j] += 1.0

    observed_disagreement = 0.0
    expected_disagreement = 0.0
    denom = float((k - 1) ** 2)

    for i in range(k):
        for j in range(k):
            weight = ((i - j) ** 2) / denom
            observed_disagreement += weight * (obs[i][j] / n)
            expected = (a_counts[i] / n) * (b_counts[j] / n)
            expected_disagreement += weight * expected

    if math.isclose(expected_disagreement, 0.0, abs_tol=1e-12):
        return 1.0 if math.isclose(observed_disagreement, 0.0, abs_tol=1e-12) else None

    return 1.0 - observed_disagreement / expected_disagreement


def pct_agreement(pairs) -> float | None:
    if not pairs:
        return None
    return sum(1 for a, b in pairs if a == b) / len(pairs)


def mean_jaccard(set_pairs) -> float | None:
    if not set_pairs:
        return None
    scores = []
    for a, b in set_pairs:
        union = a | b
        if not union:
            scores.append(1.0)
        else:
            scores.append(len(a & b) / len(union))
    return mean(scores)


def mean_set_f1(set_pairs) -> float | None:
    if not set_pairs:
        return None
    scores = []
    for a, b in set_pairs:
        if not a and not b:
            scores.append(1.0)
            continue
        denom = len(a) + len(b)
        scores.append((2.0 * len(a & b) / denom) if denom else 1.0)
    return mean(scores)


def macro_labelwise_kappa(set_pairs) -> tuple[float | None, int]:
    """
    Mean one-vs-rest Cohen kappa over substantive labels observed in either coder.

    IE/NA are deliberately excluded from the label universe. Disagreement where one
    coder uses IE/NA and the other uses substantive labels is still reflected because
    the substantive labels are absent for the IE/NA coder.
    """
    labels = sorted({
        label
        for a, b in set_pairs
        for label in (a | b)
        if label not in {"IE", "NA", ""}
    })

    kappas = []
    for label in labels:
        binary_pairs = [
            (label in a, label in b)
            for a, b in set_pairs
        ]
        k = cohen_kappa(binary_pairs)
        if k is not None:
            kappas.append(k)

    return (mean(kappas) if kappas else None, len(kappas))


def field_agreement(records_a, records_b, field: str) -> dict:
    """
    Compute field-specific reliability according to field type.
    IE/NA remain categories for unweighted nominal/set agreement.
    """
    by_a = {case_id(r["Case ID"]): r for r in records_a}
    by_b = {case_id(r["Case ID"]): r for r in records_b}
    shared = sorted(set(by_a) & set(by_b), key=lambda x: (len(x), x))

    base = {
        "field": field,
        "n_paired": len(shared),
        "annotator_c_ie_na": sum(
            1 for cid in shared if is_missing_code(by_a[cid].get(field))
        ),
        "annotator_d_ie_na": sum(
            1 for cid in shared if is_missing_code(by_b[cid].get(field))
        ),
        "annotator_c_blank": sum(
            1 for cid in shared if is_blank(by_a[cid].get(field))
        ),
        "annotator_d_blank": sum(
            1 for cid in shared if is_blank(by_b[cid].get(field))
        ),
    }

    if field in MULTILABEL_FIELDS:
        pairs = [
            (split_multilabel(field, by_a[cid].get(field)),
             split_multilabel(field, by_b[cid].get(field)))
            for cid in shared
        ]
        exact = pct_agreement(pairs)
        macro_k, n_kappas = macro_labelwise_kappa(pairs)
        base.update({
            "type": "multilabel",
            "exact_agreement": exact,
            "cohen_kappa": None,
            "quadratic_weighted_kappa": None,
            "n_ordinal_pairs": None,
            "mean_jaccard": mean_jaccard(pairs),
            "mean_set_f1": mean_set_f1(pairs),
            "macro_labelwise_kappa": macro_k,
            "n_labelwise_kappas": n_kappas,
        })
        return base

    pairs = [
        (normalize_value(field, by_a[cid].get(field)),
         normalize_value(field, by_b[cid].get(field)))
        for cid in shared
    ]
    exact = pct_agreement(pairs)
    kappa = cohen_kappa(pairs)

    if field in ORDINAL_FIELDS:
        allowed = set(ORDINAL_FIELDS[field])
        ordinal_pairs = [(a, b) for a, b in pairs if a in allowed and b in allowed]
        qwk = quadratic_weighted_kappa(ordinal_pairs, ORDINAL_FIELDS[field])
        base.update({
            "type": "ordinal",
            "exact_agreement": exact,
            "cohen_kappa": kappa,
            "quadratic_weighted_kappa": qwk,
            "n_ordinal_pairs": len(ordinal_pairs),
            "mean_jaccard": None,
            "mean_set_f1": None,
            "macro_labelwise_kappa": None,
            "n_labelwise_kappas": None,
        })
        return base

    base.update({
        "type": "nominal",
        "exact_agreement": exact,
        "cohen_kappa": kappa,
        "quadratic_weighted_kappa": None,
        "n_ordinal_pairs": None,
        "mean_jaccard": None,
        "mean_set_f1": None,
        "macro_labelwise_kappa": None,
        "n_labelwise_kappas": None,
    })
    return base


def missingness_agreement(headers, records_a, records_b) -> list[dict]:
    """
    For all 32 substantive fields, compare only status:
        coded vs IE/NA vs blank.
    This allows structured text fields such as Missing Safeguard to be assessed
    for evidence-availability agreement without pretending their wording is nominal.
    """
    fields = substantive_fields(headers)
    by_a = {case_id(r["Case ID"]): r for r in records_a}
    by_b = {case_id(r["Case ID"]): r for r in records_b}
    shared = sorted(set(by_a) & set(by_b), key=lambda x: (len(x), x))

    out = []
    for field in fields:
        pairs = [
            (evidence_status(by_a[cid].get(field)),
             evidence_status(by_b[cid].get(field)))
            for cid in shared
        ]
        out.append({
            "field": field,
            "n_paired": len(pairs),
            "c_ie_na": sum(1 for a, _ in pairs if a == "IE/NA"),
            "d_ie_na": sum(1 for _, b in pairs if b == "IE/NA"),
            "c_blank": sum(1 for a, _ in pairs if a == "blank"),
            "d_blank": sum(1 for _, b in pairs if b == "blank"),
            "status_exact_agreement": pct_agreement(pairs),
            "status_kappa": cohen_kappa(pairs),
        })
    return out


def disagreement_rows(records_a, records_b, fields) -> list[dict]:
    by_a = {case_id(r["Case ID"]): r for r in records_a}
    by_b = {case_id(r["Case ID"]): r for r in records_b}
    shared = sorted(set(by_a) & set(by_b), key=lambda x: (len(x), x))

    out = []
    for cid in shared:
        for field in fields:
            if field in MULTILABEL_FIELDS:
                a = split_multilabel(field, by_a[cid].get(field))
                b = split_multilabel(field, by_b[cid].get(field))
                equal = a == b
                a_show = "; ".join(sorted(a))
                b_show = "; ".join(sorted(b))
            else:
                a_show = normalize_value(field, by_a[cid].get(field))
                b_show = normalize_value(field, by_b[cid].get(field))
                equal = a_show == b_show

            if not equal:
                out.append({
                    "case_id": cid,
                    "field": field,
                    "annotator_c": a_show,
                    "annotator_d": b_show,
                })
    return out


def confusion_rows(records_a, records_b, fields) -> list[dict]:
    """Long-form confusion counts for nominal and ordinal fields."""
    by_a = {case_id(r["Case ID"]): r for r in records_a}
    by_b = {case_id(r["Case ID"]): r for r in records_b}
    shared = sorted(set(by_a) & set(by_b), key=lambda x: (len(x), x))

    out = []
    for field in fields:
        if field in MULTILABEL_FIELDS:
            continue
        counts = Counter(
            (normalize_value(field, by_a[cid].get(field)),
             normalize_value(field, by_b[cid].get(field)))
            for cid in shared
        )
        for (a, b), n in sorted(counts.items()):
            out.append({
                "field": field,
                "annotator_c": a,
                "annotator_d": b,
                "count": n,
            })
    return out


# ---------------------------------------------------------------------------
# Pairing / subset logic
# ---------------------------------------------------------------------------

def align_and_validate(records_c, records_d):
    ids_c = [case_id(r["Case ID"]) for r in records_c]
    ids_d = [case_id(r["Case ID"]) for r in records_d]

    if len(ids_c) != len(set(ids_c)):
        raise ValueError("Annotator C workbook contains duplicate Case IDs.")
    if len(ids_d) != len(set(ids_d)):
        raise ValueError("Annotator D workbook contains duplicate Case IDs.")

    set_c, set_d = set(ids_c), set(ids_d)
    if set_c != set_d:
        only_c = sorted(set_c - set_d)
        only_d = sorted(set_d - set_c)
        raise ValueError(
            f"Case-ID mismatch. Only C={only_c}; only D={only_d}"
        )

    by_c = {case_id(r["Case ID"]): r for r in records_c}
    by_d = {case_id(r["Case ID"]): r for r in records_d}

    source_mismatches = []
    for cid in sorted(set_c, key=lambda x: (len(x), x)):
        sc = raw_text(by_c[cid].get("Source Link(s) / Citation(s)"))
        sd = raw_text(by_d[cid].get("Source Link(s) / Citation(s)"))
        if sc != sd:
            source_mismatches.append(cid)

    return sorted(set_c, key=lambda x: (len(x), x)), source_mismatches


def analyze_subset(headers, records_c, records_d, name: str) -> dict:
    metrics = [
        field_agreement(records_c, records_d, field)
        for field in FORMAL_FIELDS
    ]
    return {
        "name": name,
        "n_cases": len(records_c),
        "field_agreement": metrics,
        "missingness_agreement": missingness_agreement(headers, records_c, records_d),
        "disagreements": disagreement_rows(records_c, records_d, FORMAL_FIELDS),
        "confusions": confusion_rows(
            records_c, records_d, NOMINAL_FIELDS + list(ORDINAL_FIELDS)
        ),
    }


# ---------------------------------------------------------------------------
# Reporting helpers
# ---------------------------------------------------------------------------

def fmt_pct(value) -> str:
    return "n/a" if value is None else f"{100.0 * value:.1f}%"


def fmt_num(value) -> str:
    return "n/a" if value is None else f"{value:.3f}"


def metric_lookup(subset, field):
    return next(m for m in subset["field_agreement"] if m["field"] == field)


def print_field_metric(m):
    field = m["field"]
    exact = fmt_pct(m["exact_agreement"])

    if m["type"] == "ordinal":
        print(
            f"  {field:<34} exact={exact:<7} "
            f"kappa={fmt_num(m['cohen_kappa']):<6} "
            f"kappa_qw={fmt_num(m['quadratic_weighted_kappa']):<6} "
            f"n_qw={m['n_ordinal_pairs']}"
        )
    elif m["type"] == "multilabel":
        print(
            f"  {field:<34} exact={exact:<7} "
            f"Jaccard={fmt_num(m['mean_jaccard']):<6} "
            f"macro_kappa={fmt_num(m['macro_labelwise_kappa']):<6}"
        )
    else:
        print(
            f"  {field:<34} exact={exact:<7} "
            f"kappa={fmt_num(m['cohen_kappa'])}"
        )


def print_report(result):
    print("=" * 86)
    print("AIFTax held-out inter-rater agreement analysis")
    print("=" * 86)
    print(f"Annotator C : {result['source_files']['annotator_c']}")
    print(f"Annotator D : {result['source_files']['annotator_d']}")
    print(f"Paired held-out cases : {result['heldout']['n_cases']}")

    if result["source_link_mismatch_case_ids"]:
        print(
            "WARNING - source-link text differs for case IDs: "
            f"{result['source_link_mismatch_case_ids']}"
        )
        print("Agreement is still paired by Case ID.")
    else:
        print("Administrative source-link fields: identical across annotators")

    for key in ("heldout",):
        subset = result[key]

        print("\n" + "-" * 86)
        print(f"{subset['name']} (n={subset['n_cases']})")
        print("-" * 86)

        print("Principal analytical fields:")
        for field in PRINCIPAL_FIELDS:
            print_field_metric(metric_lookup(subset, field))

        print("\nOther controlled fields:")
        for m in subset["field_agreement"]:
            if m["field"] not in PRINCIPAL_FIELDS:
                print_field_metric(m)

        print(
            f"\nFormal-field disagreements: {len(subset['disagreements'])} "
            f"cell(s) across {len(FORMAL_FIELDS)} controlled fields."
        )

        # Recovery missingness diagnostic.
        rec = next(
            x for x in subset["missingness_agreement"]
            if x["field"] == "Recovery Evidence Status"
        )
        print(
            "Recovery Evidence Status missingness: "
            f"C IE/NA={rec['c_ie_na']}, D IE/NA={rec['d_ie_na']}, "
            f"status agreement={fmt_pct(rec['status_exact_agreement'])}, "
            f"status kappa={fmt_num(rec['status_kappa'])}"
        )

    print("\nInterpretation:")
    print("  * Cohen kappa is unweighted and includes IE/NA as explicit categories.")
    print("  * Ordinal kappa_qw excludes pairs containing IE/NA/blank.")
    print("  * Multi-label fields report exact-set agreement, mean Jaccard, and")
    print("    macro one-vs-rest labelwise Cohen kappa.")
    print("  * Free-text fields are not assigned a content kappa; their evidence")
    print("    availability is assessed in the missingness-agreement output.")
    print("  * Recovery Evidence Status appears twice with different values: the")
    print("    field kappa above treats Documented/Inferred/IE as three categories,")
    print("    while the missingness kappa collapses them to coded vs IE/NA.")


def write_csv(path: Path, rows: list[dict], fieldnames=None):
    path.parent.mkdir(parents=True, exist_ok=True)
    if fieldnames is None:
        fieldnames = list(rows[0].keys()) if rows else []
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        if rows:
            writer.writerows(rows)


def write_outputs(result, output_dir: Path, prefix: str):
    output_dir.mkdir(parents=True, exist_ok=True)

    json_path = output_dir / f"{prefix}_report.json"
    json_path.write_text(
        json.dumps(result, indent=2, ensure_ascii=False),
        encoding="utf-8",
    )

    common_fields = [
        "field", "type", "n_paired",
        "annotator_c_ie_na", "annotator_d_ie_na",
        "annotator_c_blank", "annotator_d_blank",
        "exact_agreement", "cohen_kappa",
        "quadratic_weighted_kappa", "n_ordinal_pairs",
        "mean_jaccard", "mean_set_f1",
        "macro_labelwise_kappa", "n_labelwise_kappas",
    ]

    for key in ("heldout",):
        subset = result[key]

        write_csv(
            output_dir / f"{prefix}_field_agreement.csv",
            subset["field_agreement"],
            common_fields,
        )
        write_csv(
            output_dir / f"{prefix}_missingness_agreement.csv",
            subset["missingness_agreement"],
        )
        write_csv(
            output_dir / f"{prefix}_disagreements.csv",
            subset["disagreements"],
            ["case_id", "field", "annotator_c", "annotator_d"],
        )
        write_csv(
            output_dir / f"{prefix}_confusions.csv",
            subset["confusions"],
            ["field", "annotator_c", "annotator_d", "count"],
        )

    print(f"\nJSON report            : {json_path}")
    print(
        f"Field agreement        : "
        f"{output_dir / f'{prefix}_field_agreement.csv'}"
    )
    print(
        f"Missingness agreement  : "
        f"{output_dir / f'{prefix}_missingness_agreement.csv'}"
    )
    print(
        f"Disagreements          : "
        f"{output_dir / f'{prefix}_disagreements.csv'}"
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description=(
            "Compute field-specific inter-rater agreement between two independent "
            "AIFTax held-out annotators."
        )
    )
    parser.add_argument(
        "--annotator-c", required=True,
        help="Completed Annotator C workbook."
    )
    parser.add_argument(
        "--annotator-d", required=True,
        help="Completed Annotator D workbook."
    )
    parser.add_argument(
        "--output-dir", default="results_heldout_agreement"
    )
    parser.add_argument(
        "--output-prefix", default="AIFTax_heldout"
    )
    args = parser.parse_args()

    c_path = Path(args.annotator_c)
    d_path = Path(args.annotator_d)

    headers_c, c_records = load_records(c_path)
    headers_d, d_records = load_records(d_path)

    # Check required formal fields exist after stripping header whitespace.
    missing_c = [f for f in FORMAL_FIELDS if f not in headers_c]
    missing_d = [f for f in FORMAL_FIELDS if f not in headers_d]
    if missing_c or missing_d:
        raise ValueError(
            f"Missing agreement fields. Annotator C={missing_c}; Annotator D={missing_d}"
        )

    shared_ids, source_mismatches = align_and_validate(c_records, d_records)

    heldout = analyze_subset(headers_c, c_records, d_records, "Held-out set")

    result = {
        "source_files": {
            "annotator_c": c_path.name,
            "annotator_d": d_path.name,
        },
        "paired_case_ids": shared_ids,
        "source_link_mismatch_case_ids": source_mismatches,
        "method_notes": {
            "denominator": (
                "All cases independently annotated by both C and D; no case is "
                "dropped on the basis of an annotation outcome."
            ),
            "unweighted_kappa": (
                "IE/NA retained as explicit categories."
            ),
            "ordinal_weighted_kappa": (
                "Quadratic-weighted kappa calculated only when both annotations "
                "have a valid ordinal value; IE/NA/blank excluded."
            ),
            "multilabel": (
                "Exact set agreement, mean Jaccard, mean set F1, and macro "
                "one-vs-rest labelwise Cohen kappa."
            ),
            "free_text": (
                "No content kappa; only coded-vs-IE/NA-vs-blank status agreement."
            ),
        },
        "heldout": heldout,
    }

    print_report(result)
    write_outputs(result, Path(args.output_dir), args.output_prefix)


if __name__ == "__main__":
    main()
