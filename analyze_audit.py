#!/usr/bin/env python3
"""
AIFTAX author-coded validation audit: analysis.

    python analyze_audit.py \
        --audit "results_audit_draw/AIFTax_Audit_Coding_Workbook_Completed.xlsx" \
        --adjudicated "Final_Corpus_and_Adjudicated_Log/AIFTax_Final_Adjudicated_Dataset.xlsx" \
        --annotator-a "Annotation Materials/AIFTax_Annotator_A_Annotation_Completed.xlsx" \
        --annotator-b "Annotation Materials/AIFTax_Annotator_B_Annotation_Completed.xlsx" \
        --output-dir results_audit_analysis

The audit is scored against the FINAL ADJUDICATED dataset, because that is the artifact
the paper's findings rest on. The A/B workbooks supply cell-level provenance only:
whether each adjudicated label came from LLM consensus (A=B) or an author decision (A!=B).
"""

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from math import comb
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT_MARKERS = ("Final_Corpus_and_Adjudicated_Log", "Annotation Materials", ".git")
TABLE_V = {"Failure Category": 11, "Propagation Reach": 36,
           "Risk Assessment": 17, "Recovery Complexity": 9}
PRINCIPAL = ["Failure Category", "Propagation Reach", "Risk Assessment", "Recovery Complexity"]

ORDINAL_SCALES = {
    "Propagation Reach":   ["p0", "p1", "p2", "p3"],
    "Risk Assessment":     ["low", "moderate", "high", "severe", "critical"],
    "Recovery Complexity": ["low", "moderate", "high", "severe", "very high"],
}
HIGH_PLUS_FROM = {"Risk Assessment": 2, "Recovery Complexity": 2}   # index of "high"

MULTI_LABEL = {"Failure Mode(s)", "Impact Dimension(s)",
               "Causal Location(s)", "Recovery / Maintenance Action(s)"}
FREE_TEXT = {"Date", "Entity", "Title", "Application Domain", "Timing / Period",
             "Manifestation — Brief Description", "Propagation Chain", "Risk Subdomain",
             "Causal Mechanism", "Missing Safeguard", "Root Cause Description"}

ARIAL = "Arial"
HDR = PatternFill("solid", fgColor="D9D9D9")
WARN = PatternFill("solid", fgColor="FFE0E0")
THIN = Side(style="thin", color="BFBFBF")


def clean(s):
    return re.sub(r"\s+", " ", str(s)).replace("\u00a0", " ").strip()


def nrm(s):
    """Normalise a scalar cell value for comparison."""
    if isinstance(s, (pd.Timestamp, datetime)):
        return s.strftime("%Y-%m-%d")
    t = clean(s)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2} 00:00:00", t):
        t = t[:10]
    return t.casefold()


def sset(s):
    return frozenset(x.strip() for x in nrm(s).split(";") if x.strip())


def cmp_for(field):
    return sset if field in MULTI_LABEL else nrm


def find_root(explicit=None):
    if explicit:
        return Path(explicit).expanduser().resolve()
    for start in (Path.cwd(), Path(__file__).resolve().parent):
        for d in (start, *start.parents):
            if any((d / m).exists() for m in ROOT_MARKERS):
                return d
    return Path.cwd()


def rin(raw, root, label):
    p = Path(raw).expanduser()
    tried = [p] if p.is_absolute() else list(dict.fromkeys([Path.cwd() / p, root / p]))
    for t in tried:
        if t.exists():
            return t.resolve()
    print(f"ERROR: {label} not found. Tried:", file=sys.stderr)
    for t in tried:
        print(f"    {t}", file=sys.stderr)
    sys.exit(1)


def detect_header(path, sheet, max_scan=12):
    raw = pd.read_excel(path, sheet_name=sheet, header=None, nrows=max_scan)
    for i in range(len(raw)):
        if "case id" in [clean(v).lower() for v in raw.iloc[i].tolist() if pd.notna(v)]:
            return i
    return 0


def load_wide(path, prefer=("Final Corpus", "Annotations", "Audit")):
    xl = pd.ExcelFile(path)
    best, bestscore = None, -1
    for s in xl.sheet_names:
        hdr = detect_header(path, s)
        d = xl.parse(s, header=hdr)
        cols = [clean(c) for c in d.columns]
        if "case id" not in [c.lower() for c in cols]:
            continue
        score = (100 if s in prefer else 0) + (50 if 20 <= len(d.dropna(how="all")) <= 200 else 0) \
                + (40 if abs(len(cols) - 34) <= 3 else 0)
        if any(c.lower() in ("annotator a value", "final value") for c in cols):
            score -= 300
        if score > bestscore:
            best, bestscore, besthdr = s, score, hdr
    if best is None:
        sys.exit(f"No wide sheet with a Case ID column in {path.name}")
    d = xl.parse(best, header=besthdr)
    d.columns = [clean(c) for c in d.columns]
    d = d.dropna(how="all")
    cid = next(c for c in d.columns if c.lower() == "case id")
    d = d[d[cid].notna()].copy()
    d[cid] = d[cid].astype(int)
    return d.set_index(cid), best


def kappa(x, y, weighted=False):
    labs = sorted(set(x) | set(y), key=str)
    if len(labs) < 2:
        return float("nan")
    idx = {l: i for i, l in enumerate(labs)}
    m = len(x)
    O = np.zeros((len(labs), len(labs)))
    for i, j in zip(x, y):
        O[idx[i], idx[j]] += 1
    O /= m
    E = np.outer(O.sum(1), O.sum(0))
    if weighted:
        k = len(labs) - 1
        W = np.array([[(i - j) ** 2 / k ** 2 for j in range(len(labs))] for i in range(len(labs))])
    else:
        W = 1 - np.eye(len(labs))
    den = (W * E).sum()
    return float("nan") if den == 0 else 1 - (W * O).sum() / den


def boot_ci(x, y, weighted=False, n=2000, seed=20260830):
    rng = np.random.default_rng(seed)
    m = len(x)
    vals = []
    for _ in range(n):
        s = rng.integers(0, m, m)
        v = kappa([x[i] for i in s], [y[i] for i in s], weighted)
        if not np.isnan(v):
            vals.append(v)
    return (round(float(np.percentile(vals, 2.5)), 3),
            round(float(np.percentile(vals, 97.5)), 3)) if vals else (None, None)


def sign_p(lo, hi):
    nd = lo + hi
    if nd == 0:
        return 1.0
    k = max(lo, hi)
    return min(1.0, 2 * sum(comb(nd, i) for i in range(k, nd + 1)) / 2 ** nd)


def level(v, scale):
    s = nrm(v).split("—")[0].split(" - ")[0].strip()
    for i, x in enumerate(scale):
        if s.startswith(x):
            return i
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--audit", required=True)
    ap.add_argument("--adjudicated", required=True)
    ap.add_argument("--annotator-a", required=True)
    ap.add_argument("--annotator-b", required=True)
    ap.add_argument("--output-dir", default="results_audit_analysis")
    ap.add_argument("--project-root", default=None)
    args = ap.parse_args()

    root = find_root(args.project_root)
    p_aud = rin(args.audit, root, "--audit")
    p_adj = rin(args.adjudicated, root, "--adjudicated")
    p_a = rin(args.annotator_a, root, "--annotator-a")
    p_b = rin(args.annotator_b, root, "--annotator-b")
    out = Path(args.output_dir)
    out = out.resolve() if out.is_absolute() else (root / out).resolve()
    out.mkdir(parents=True, exist_ok=True)

    corp, s_corp = load_wide(p_adj)
    aud, s_aud = load_wide(p_aud)
    A, s_a = load_wide(p_a)
    B, s_b = load_wide(p_b)
    log = pd.ExcelFile(p_adj).parse("Adjudication Log", header=0)
    log.columns = [clean(c) for c in log.columns]

    src_col = next(c for c in corp.columns if "source link" in c.lower() or "citation" in c.lower())
    fields = [c for c in corp.columns if c != src_col]
    shared = [f for f in fields if f in A.columns and f in B.columns]
    cases = sorted(aud.index)

    print(f"Corpus     : {p_adj.name} ['{s_corp}']  {len(corp)} cases, {len(fields)} fields")
    print(f"Audit      : {p_aud.name} ['{s_aud}']  {len(cases)} cases")
    print(f"Annotator A: {p_a.name} ['{s_a}']   B: {p_b.name} ['{s_b}']   shared fields: {len(shared)}")

    # ================= A. INPUT VERIFICATION ==================================
    print("\n" + "=" * 78 + "\nA. INPUT VERIFICATION\n" + "=" * 78)
    ver = []
    common = [i for i in corp.index if i in A.index and i in B.index]

    for var, exp in TABLE_V.items():
        c = cmp_for(var)
        got = sum(1 for i in common if c(A.at[i, var]) != c(B.at[i, var]))
        ok = got == exp
        ver.append((f"Table V Diff. — {var}", got, exp, "ok" if ok else "MISMATCH"))
        print(f"  {var:<22} A!=B on {got:>4}   paper reports {exp:<4} {'ok' if ok else 'MISMATCH'}")

    ab_keys = {(i, f) for i in common for f in shared
               if cmp_for(f)(A.at[i, f]) != cmp_for(f)(B.at[i, f])}
    log_keys = {(int(r["Case ID"]), clean(r["Field"])) for _, r in log.iterrows()}
    only_ab = ab_keys - log_keys
    only_log = log_keys - ab_keys
    print(f"\n  A!=B cells (set-aware)      : {len(ab_keys)}")
    print(f"  Adjudication-log keys        : {len(log_keys)}")
    print(f"  In A!=B but not logged       : {len(only_ab)}")
    print(f"  Logged but A==B              : {len(only_log)}")
    ver += [("A!=B cells (set-aware)", len(ab_keys), len(log_keys), "ok" if not only_ab and not only_log else "MISMATCH"),
            ("Disagreements not in log", len(only_ab), 0, "ok" if not only_ab else "REVIEW"),
            ("Logged entries with A==B", len(only_log), 0, "ok" if not only_log else "REVIEW")]
    for k in sorted(only_ab)[:5]:
        print(f"      unlogged: {k}")
    for k in sorted(only_log)[:5]:
        print(f"      spurious: {k}")

    # consensus sweep: A==B but corpus differs from both
    consensus_exceptions = []
    n_consensus = 0
    for i in common:
        for f in shared:
            c = cmp_for(f)
            if c(A.at[i, f]) == c(B.at[i, f]):
                n_consensus += 1
                if c(corp.at[i, f]) != c(A.at[i, f]):
                    consensus_exceptions.append(
                        {"Case ID": i, "Field": f, "LLM A = LLM B": clean(A.at[i, f]),
                         "Final Corpus": clean(corp.at[i, f])})
    print(f"\n  Consensus cells (A==B)       : {n_consensus}")
    print(f"  Consensus cells where the final corpus differs from BOTH models: "
          f"{len(consensus_exceptions)}")
    for e in consensus_exceptions[:10]:
        print(f"      case {e['Case ID']:>3}  {e['Field']:<28} models={e['LLM A = LLM B'][:26]!r} "
              f"corpus={e['Final Corpus'][:26]!r}")
    ver += [("Consensus cells (A==B)", n_consensus, 3200 - len(log_keys), ""),
            ("Consensus cells corrected in corpus", len(consensus_exceptions), 0,
             "ok" if not consensus_exceptions else "DOCUMENT")]

    # ================= B. AUDIT SCORING =======================================
    print("\n" + "=" * 78 + "\nB. AUDIT vs FINAL ADJUDICATED DATASET\n" + "=" * 78)
    prov = {(i, f): ("adjudicated" if (i, f) in log_keys else "consensus")
            for i in cases for f in fields}

    rows = []
    for f in fields:
        c = cmp_for(f)
        kind = ("Multi-label" if f in MULTI_LABEL
                else "Free text / narrative" if f in FREE_TEXT else "Controlled")
        hits = [(i, c(aud.at[i, f]) == c(corp.at[i, f])) for i in cases]
        n = len(hits)
        ex = sum(h for _, h in hits)
        cn = [(i, h) for i, h in hits if prov[(i, f)] == "consensus"]
        ad = [(i, h) for i, h in hits if prov[(i, f)] == "adjudicated"]
        r = {"Field": f, "Kind": kind, "N": n, "Exact %": round(100 * ex / n, 1),
             "Consensus n": len(cn),
             "Consensus %": round(100 * sum(h for _, h in cn) / len(cn), 1) if cn else None,
             "Adjudicated n": len(ad),
             "Adjudicated %": round(100 * sum(h for _, h in ad) / len(ad), 1) if ad else None}
        if kind == "Controlled":
            x = [nrm(aud.at[i, f]) for i in cases]
            y = [nrm(corp.at[i, f]) for i in cases]
            r["kappa"] = round(kappa(x, y), 3)
            if f in ORDINAL_SCALES:
                sc = ORDINAL_SCALES[f]
                xl_ = [level(aud.at[i, f], sc) for i in cases]
                yl = [level(corp.at[i, f], sc) for i in cases]
                pairs = [(a, b) for a, b in zip(xl_, yl) if a is not None and b is not None]
                if pairs:
                    xa, ya = zip(*pairs)
                    r["kappa_qw"] = round(kappa(list(xa), list(ya), weighted=True), 3)
                    d = [abs(a - b) for a, b in zip(xa, ya)]
                    r["Adjacent %"] = round(100 * sum(1 for k in d if k <= 1) / len(d), 1)
                    lo = sum(1 for a, b in zip(xa, ya) if a < b)
                    hi = sum(1 for a, b in zip(xa, ya) if a > b)
                    r["Lower"], r["Higher"] = lo, hi
                    r["Sign p"] = float(f"{sign_p(lo, hi):.2e}")
        elif kind == "Multi-label":
            J, tp, fp, fn = [], 0, 0, 0
            labels = set()
            for i in cases:
                sa, sb = sset(aud.at[i, f]), sset(corp.at[i, f])
                labels |= sa | sb
                J.append(len(sa & sb) / len(sa | sb) if (sa | sb) else 1.0)
                tp += len(sa & sb); fp += len(sa - sb); fn += len(sb - sa)
            r["Mean Jaccard"] = round(100 * float(np.mean(J)), 1)
            r["Micro F1"] = round(2 * tp / (2 * tp + fp + fn), 3) if tp else 0.0
            ks = []
            for lab in labels:
                x = [lab in sset(aud.at[i, f]) for i in cases]
                y = [lab in sset(corp.at[i, f]) for i in cases]
                k = kappa(x, y)
                if not np.isnan(k):
                    ks.append(k)
            r["Macro per-label kappa"] = round(float(np.mean(ks)), 3) if ks else None
        rows.append(r)
    res = pd.DataFrame(rows)

    print("\nPRINCIPAL VARIABLES")
    cols = ["Field", "N", "Exact %", "kappa", "kappa_qw", "Adjacent %", "Lower", "Higher",
            "Sign p", "Consensus n", "Consensus %", "Adjudicated n", "Adjudicated %"]
    pv = res[res.Field.isin(PRINCIPAL)].reindex(columns=[c for c in cols if c in res.columns])
    print(pv.to_string(index=False))

    # bootstrap CIs + High-or-worse dichotomy for the principal variables
    detail = []
    for f in PRINCIPAL:
        x = [nrm(aud.at[i, f]) for i in cases]
        y = [nrm(corp.at[i, f]) for i in cases]
        lo, hi = boot_ci(x, y)
        d = {"Variable": f, "kappa 95% CI": f"[{lo}, {hi}]"}
        if f in HIGH_PLUS_FROM:
            sc, thr = ORDINAL_SCALES[f], HIGH_PLUS_FROM[f]
            xa = [level(aud.at[i, f], sc) for i in cases]
            ya = [level(corp.at[i, f], sc) for i in cases]
            ah = sum(1 for v in xa if v is not None and v >= thr)
            ch = sum(1 for v in ya if v is not None and v >= thr)
            agr = sum(1 for a, b in zip(xa, ya) if (a >= thr) == (b >= thr))
            d.update({"High+ audit": f"{ah}/{len(cases)} ({round(100*ah/len(cases))}%)",
                      "High+ corpus": f"{ch}/{len(cases)} ({round(100*ch/len(cases))}%)",
                      "Dichotomy agreement": f"{round(100*agr/len(cases),1)}%"})
        detail.append(d)
    detail = pd.DataFrame(detail)
    print("\n" + detail.to_string(index=False))

    print("\nCONTROLLED + MULTI-LABEL, all fields (sorted by exact agreement)")
    show = res[res.Kind != "Free text / narrative"].sort_values("Exact %")
    print(show[["Field", "Kind", "Exact %", "kappa", "Consensus n", "Consensus %",
                "Adjudicated n", "Adjudicated %"]].to_string(index=False))

    print("\nMULTI-LABEL DETAIL")
    ml = res[res.Kind == "Multi-label"]
    print(ml[["Field", "Exact %", "Mean Jaccard", "Micro F1",
              "Macro per-label kappa"]].to_string(index=False))

    print("\nFREE TEXT / NARRATIVE — diagnostic only, not interpreted as disagreement")
    print(res[res.Kind == "Free text / narrative"][["Field", "Exact %"]].to_string(index=False))

    # ---- provenance rollup --------------------------------------------------
    cn_t = cn_a = ad_t = ad_a = 0
    for f in fields:
        if f in FREE_TEXT:
            continue
        c = cmp_for(f)
        for i in cases:
            h = c(aud.at[i, f]) == c(corp.at[i, f])
            if prov[(i, f)] == "consensus":
                cn_t += 1; cn_a += h
            else:
                ad_t += 1; ad_a += h
    print(f"\nPROVENANCE ROLLUP (controlled + multi-label cells only)")
    print(f"  consensus   {cn_a}/{cn_t} = {100*cn_a/cn_t:.1f}%")
    print(f"  adjudicated {ad_a}/{ad_t} = {100*ad_a/ad_t:.1f}%")

    # ---- QC: vocabulary violations in the audit -----------------------------
    lists = pd.ExcelFile(p_adj).parse("Lists", header=0)
    vocab = {clean(c): {nrm(v) for v in lists[c].dropna()} for c in lists.columns}
    alias = {"Failure Mode(s)": "Failure Mode", "Impact Dimension(s)": "Impact Class",
             "Causal Location(s)": "Causal Location",
             "Recovery / Maintenance Action(s)": "Recovery / Maintenance Actions"}
    viol = []
    for f in fields:
        key = alias.get(f, f)
        if key not in vocab:
            continue
        for i in cases:
            vals = sset(aud.at[i, f]) if f in MULTI_LABEL else {nrm(aud.at[i, f])}
            for v in vals:
                if v and v not in vocab[key] and v not in ("ie", "na", "ie/na"):
                    viol.append({"Case ID": i, "Field": f, "Audit value": clean(aud.at[i, f]),
                                 "Corpus value": clean(corp.at[i, f])})
                    break
    print(f"\nQC: controlled-vocabulary violations in the audit workbook: {len(viol)}")
    for v in viol:
        print(f"      case {v['Case ID']:>3}  {v['Field']}: {v['Audit value']!r} "
              f"(corpus {v['Corpus value']!r})")
    print(f"QC: IE/NA codes used by the author coder: "
          f"{sum(1 for f in fields for i in cases if nrm(aud.at[i,f]) in ('ie','na','ie/na'))}"
          f" of {len(cases)*len(fields)}")

    # ================= C. WRITE ==============================================
    wb = Workbook()
    ws = wb.active; ws.title = "README"
    lines = [
        ("AIFTAX — author-coded validation audit: analysis", True), ("", False),
        ("The audit is scored against the FINAL ADJUDICATED dataset, the artifact the paper's", False),
        ("findings rest on. The pre-adjudication workbooks supply cell-level provenance only:", False),
        ("whether each adjudicated label came from LLM consensus (A=B) or an author decision.", False),
        ("", False),
        ("The coder is an author of the study, withheld from the adjudicated labels, the model", False),
        ("outputs and the key, but not an independent rater. Agreement measures reproducibility", False),
        ("under the instrument, not inter-rater reliability.", False), ("", False),
        ("Multi-label values are compared as unordered sets. Free-text and narrative fields are", False),
        ("reported as a diagnostic only; wording differences are not categorical disagreement.", False),
        ("Ordinal fields additionally report quadratic-weighted kappa, adjacency, the direction of", False),
        ("disagreement and a two-sided sign test. Quadratic weighting forgives near-misses, so it", False),
        ("must be read alongside the direction columns when disagreement is one-sided.", False),
        ("", False),
        (f"Sources: {p_adj.name}, {p_aud.name}, {p_a.name}, {p_b.name}", False),
        (f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}", False),
    ]
    for i, (t, b) in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=t).font = Font(name=ARIAL, size=11, bold=b)
    ws.column_dimensions["A"].width = 102

    def sheet(name, df, title, warn_col=None):
        w = wb.create_sheet(name)
        w.cell(row=1, column=1, value=title).font = Font(name=ARIAL, size=11, bold=True)
        for j, c in enumerate(df.columns, start=1):
            cc = w.cell(row=3, column=j, value=c)
            cc.font = Font(name=ARIAL, size=10, bold=True); cc.fill = HDR
            cc.border = Border(bottom=THIN)
            cc.alignment = Alignment(wrap_text=True, horizontal="left" if j == 1 else "right")
        for r, (_, row) in enumerate(df.iterrows(), start=4):
            for j, c in enumerate(df.columns, start=1):
                v = row[c]
                cc = w.cell(row=r, column=j, value=None if pd.isna(v) else v)
                cc.font = Font(name=ARIAL, size=10)
                cc.alignment = Alignment(horizontal="left" if j == 1 else "right")
                if warn_col and c == warn_col and str(v) not in ("ok", "", "nan"):
                    cc.fill = WARN
        for j, c in enumerate(df.columns, start=1):
            w.column_dimensions[get_column_letter(j)].width = 34 if j == 1 else 15
        w.freeze_panes = w.cell(row=4, column=2)

    sheet("Verification", pd.DataFrame(ver, columns=["Check", "Value", "Expected", "Status"]),
          "Input verification", warn_col="Status")
    sheet("Principal Variables", pv, "Audit vs adjudicated — four principal variables")
    sheet("Principal Detail", detail, "Bootstrap CIs and High-or-worse dichotomy")
    sheet("All Fields", res, "Audit vs adjudicated — all 32 substantive fields")
    if consensus_exceptions:
        sheet("Consensus Exceptions", pd.DataFrame(consensus_exceptions),
              "Consensus cells where the final corpus differs from both models")
    if viol:
        sheet("QC Violations", pd.DataFrame(viol),
              "Controlled-vocabulary violations in the audit workbook")

    xlsx = out / "AIFTax_Audit_Analysis.xlsx"
    wb.save(xlsx)
    res.to_csv(out / "audit_all_fields.csv", index=False)
    pv.to_csv(out / "audit_principal_variables.csv", index=False)
    pd.DataFrame(ver, columns=["Check", "Value", "Expected", "Status"]).to_csv(
        out / "audit_verification.csv", index=False)
    if consensus_exceptions:
        pd.DataFrame(consensus_exceptions).to_csv(out / "consensus_exceptions.csv", index=False)

    (out / "audit_analysis_summary.json").write_text(json.dumps({
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "n_cases": len(cases), "case_ids": [int(c) for c in cases],
        "table_v_reproduced": all(v[3] == "ok" for v in ver[:4]),
        "log_exhaustive": not only_ab and not only_log,
        "consensus_cells": n_consensus,
        "consensus_exceptions": consensus_exceptions,
        "vocabulary_violations": viol,
        "provenance_rollup": {"consensus_pct": round(100 * cn_a / cn_t, 1),
                              "adjudicated_pct": round(100 * ad_a / ad_t, 1)},
    }, indent=2, default=str))

    print(f"\nWrote {xlsx.name} + CSVs + summary JSON to {out}")


if __name__ == "__main__":
    main()
