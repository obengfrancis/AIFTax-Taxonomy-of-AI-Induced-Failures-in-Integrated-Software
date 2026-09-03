#!/usr/bin/env python3
"""
AIFTAX author-coded validation audit: reproducible 30-case sample draw + blinded
32-field coding workbook.

Runnable from the project root or from anywhere:

    python draw_audit_sample.py --dry-run

    python draw_audit_sample.py \
        --adjudicated "Final_Corpus_and_Adjudicated_Log/AIFTax_Final_Adjudicated_Dataset.xlsx" \
        --annotator-a "Annotation Materials/AIFTax_Annotator_A_Annotation_Completed.xlsx" \
        --annotator-b "Annotation Materials/AIFTax_Annotator_B_Annotation_Completed.xlsx" \
        --output-dir results_audit_draw

Path resolution: absolute paths are used as given; relative paths are tried against the
current directory first, then against the discovered project root. Relative --output-dir
resolves against the project root, so outputs land in the repo wherever you invoke from.

Outputs, under <output-dir>/ :
    AIFTax_Audit_Coding_Workbook.xlsx   <- the only file the coder opens
    aiftax_audit_draw_record.json       <- post-draw reproducibility record
    SEALED/aiftax_audit_KEY.xlsx        <- adjudicated labels + cell provenance; do not open

Design notes:
  - The primary audit sample is a simple random sample of 30 incidents without replacement.
    No outcome-based stratification or post-draw census cases are added.
  - All 32 substantive fields are blank in the coding workbook and are coded by the author.
  - The coding-sheet row order is shuffled independently so coding order carries no signal.
  - The visual formatting is copied directly from the Annotator A completed workbook.
    Only formatting is copied: Annotator A values are never copied into audit fields.
  - The audit is performed by an author of the study. The coder is withheld from adjudicated
    labels, model outputs, provenance flags, and the sealed key; agreement therefore measures
    reproducibility under the instrument, not independent-human inter-rater reliability.
  - Multi-label comparisons used for provenance treat label order as irrelevant.
  - The Markdown audit protocol should be committed before this script is run. The script
    records its SHA-256 digest in the draw record when --protocol-file is available.
"""

import argparse
import hashlib
import json
import random
import re
import sys
from copy import copy
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- DEFAULTS ---
DEFAULT_ADJ = "Final_Corpus_and_Adjudicated_Log/AIFTax_Final_Adjudicated_Dataset.xlsx"
DEFAULT_A   = "Annotation Materials/AIFTax_Annotator_A_Annotation_Completed.xlsx"
DEFAULT_B   = "Annotation Materials/AIFTax_Annotator_B_Annotation_Completed.xlsx"
DEFAULT_OUT = "results_audit_draw"

ROOT_MARKERS = ("Final_Corpus_and_Adjudicated_Log", "Annotation Materials", ".git")

CASE_ID_CANON = "case id"
SOURCE_HINTS  = ("source link", "citation")

N_RANDOM    = 30
SAMPLE_SEED = 20260827
ORDER_SEED  = 27082026
EXPECT_SUBSTANTIVE = 32

# Published pre-adjudication disagreement counts (Table V, n=100). Used to confirm the
# two workbooks loaded really are the unchanged pre-adjudication pair.
EXPECTED_DIFFS = {"Failure Category": 11, "Propagation Reach": 36,
                  "Risk Assessment": 17, "Recovery Complexity": 9}

MULTILABEL_FIELDS = {
    "Failure Mode(s)",
    "Impact Dimension(s)",
    "Recovery / Maintenance Action(s)",
    "Causal Location(s)",
}
# -----------------------------------------------------------------------------

ARIAL = "Arial"
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
FILLIN_FILL = PatternFill("solid", fgColor="FFF2CC")
BAND_FILL   = PatternFill("solid", fgColor="BFBFBF")


def clean(h):
    """Collapse internal whitespace; the workbooks contain e.g. 'Impact  Dimension(s)'."""
    return re.sub(r"\s+", " ", str(h)).replace("\u00a0", " ").strip()


def norm_for_compare(field, value):
    """Normalize values for A-vs-B comparison; multi-label fields are unordered."""
    if pd.isna(value):
        return ""
    v = clean(value)
    if field in MULTILABEL_FIELDS:
        labels = [clean(x).casefold() for x in v.split(";") if clean(x)]
        return tuple(sorted(labels))
    return v.casefold()


def sha256_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def find_project_root(explicit=None):
    if explicit:
        return Path(explicit).expanduser().resolve()
    for start in (Path.cwd(), Path(__file__).resolve().parent):
        for d in (start, *start.parents):
            if any((d / m).exists() for m in ROOT_MARKERS):
                return d
    return Path.cwd()


def resolve_input(raw, root, label):
    p = Path(raw).expanduser()
    candidates = [p] if p.is_absolute() else [Path.cwd() / p, root / p]
    tried = list(dict.fromkeys(candidates))   # de-duplicate when cwd == project root
    for t in tried:
        if t.exists():
            return t.resolve()
    print(f"ERROR: {label} not found. Tried:", file=sys.stderr)
    for t in tried:
        print(f"    {t}", file=sys.stderr)
    sys.exit(1)


def resolve_output(raw, root):
    p = Path(raw).expanduser()
    return p.resolve() if p.is_absolute() else (root / p).resolve()


# Column names that mark a long-format adjudication log rather than the wide corpus.
LONGFORM_MARKERS = ("annotator a value", "annotator b value", "final value",
                    "final selection", "decision rule", "adjudication rationale")


def detect_header(path, sheet, max_scan=12):
    """Find the row index holding the real header. Sheets often carry a banner row above it."""
    raw = pd.read_excel(path, sheet_name=sheet, header=None, nrows=max_scan)
    for i in range(len(raw)):
        vals = [clean(v).lower() for v in raw.iloc[i].tolist() if pd.notna(v)]
        if CASE_ID_CANON in vals:
            return i
    return 0


def survey(path):
    """Return [(sheet, header_row, n_rows, n_cols, has_case_id, score)] for every sheet."""
    out, xl = [], pd.ExcelFile(path)
    for name in xl.sheet_names:
        try:
            hdr = detect_header(path, name)
            full = xl.parse(name, header=hdr)
            cols = [clean(c) for c in full.columns]
            full = full.dropna(how="all")
            has_id = any(c.lower() == CASE_ID_CANON for c in cols)
            longform = sum(1 for c in cols if c.lower() in LONGFORM_MARKERS)
            score = (100 if has_id else 0)
            score += 50 if 50 <= len(full) <= 200 else 0
            score += 40 if abs(len(cols) - (EXPECT_SUBSTANTIVE + 2)) <= 2 else 0
            score += min(len(cols), 40)
            score -= 300 if longform >= 2 else 0          # this is the adjudication log
            out.append((name, hdr, len(full), len(cols), has_id, score))
        except Exception:
            out.append((name, -1, -1, -1, False, -999))
    return out


def load_sheet(path, sheet=None, announce=False):
    """Load the wide corpus sheet, choosing by shape unless one is named."""
    xl = pd.ExcelFile(path)
    sheets = survey(path)
    if sheet is None:
        ranked = sorted(sheets, key=lambda r: r[5], reverse=True)
        if not ranked or ranked[0][5] < 100:
            print("  Sheets:", file=sys.stderr)
            for name, hdr, nrow, ncol, has_id, score in sheets:
                print(f"    {name:<34} header_row={hdr} rows={nrow} cols={ncol} "
                      f"case_id={has_id}", file=sys.stderr)
            sys.exit(f"No sheet in {path.name} looks like the wide corpus. Pass --sheet explicitly.")
        chosen = ranked[0][0]
    else:
        chosen = sheet
    if announce:
        print("  Sheets:")
        for name, hdr, nrow, ncol, has_id, score in sheets:
            mark = "  <-- selected" if name == chosen else ""
            print(f"    {name:<24} header_row={hdr:<3} rows={nrow:<5} cols={ncol:<4} "
                  f"case_id={has_id}{mark}")
    hdr = next((s[1] for s in sheets if s[0] == chosen), 0)
    df = xl.parse(chosen, header=hdr)
    df.columns = [clean(c) for c in df.columns]
    df = df.dropna(how="all")
    cid = next((c for c in df.columns if c.lower() == CASE_ID_CANON), None)
    if cid is not None:
        df = df[df[cid].notna()]          # drop trailing notes / total rows
    return df, chosen, hdr


def read_header_bands(path, sheet, hdr_row):
    """Read the merged group band(s) sitting above the field header.

    Returns a list (one per band row) of {field name -> band label}. Merged cells carry
    their value only in the top-left cell, so ranges are expanded before mapping.
    """
    if hdr_row <= 0:
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet]
    filled = {}
    for rng in ws.merged_cells.ranges:
        v = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                filled[(r, c)] = v
    fields = {c: clean(ws.cell(row=hdr_row + 1, column=c).value)
              for c in range(1, ws.max_column + 1)
              if ws.cell(row=hdr_row + 1, column=c).value is not None}
    bands = []
    for r in range(1, hdr_row + 1):
        m = {}
        for c, fname in fields.items():
            v = filled.get((r, c), ws.cell(row=r, column=c).value)
            if v is not None:
                m[fname] = clean(v)
        bands.append(m)
    wb.close()
    return bands


def _merged_anchor(ws, row, col):
    """Return the top-left cell for a merged range containing (row, col)."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return ws.cell(rng.min_row, rng.min_col)
    return ws.cell(row, col)


def style_snapshot(cell):
    """Copy portable style components from a cell for use in another workbook."""
    return {
        "font": copy(cell.font),
        "fill": copy(cell.fill),
        "border": copy(cell.border),
        "alignment": copy(cell.alignment),
        "number_format": cell.number_format,
        "protection": copy(cell.protection),
    }


def apply_style(cell, style):
    """Apply a style snapshot to a destination cell in a different workbook."""
    if not style:
        return
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy(style["protection"])


def read_template_styles(path, sheet, hdr_row):
    """
    Read visual formatting from the completed Annotator A workbook.

    Formatting is indexed by cleaned field name. Body styles and row heights are
    additionally indexed by Case ID so row-specific banding is preserved without
    copying any Annotator A annotation values.
    """
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet]
    header_row = hdr_row + 1

    field_cols = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col).value
        if value is not None:
            field_cols[clean(value)] = col

    cid_field = next((f for f in field_cols if f.lower() == CASE_ID_CANON), None)
    case_rows = {}
    if cid_field is not None:
        cid_col = field_cols[cid_field]
        for row in range(header_row + 1, ws.max_row + 1):
            value = ws.cell(row, cid_col).value
            if value is not None:
                case_rows[value] = row

    header_styles = {f: style_snapshot(ws.cell(header_row, c))
                     for f, c in field_cols.items()}
    widths = {f: ws.column_dimensions[get_column_letter(c)].width
              for f, c in field_cols.items()}

    band_styles, band_heights = [], []
    for row in range(1, header_row):
        band_styles.append({f: style_snapshot(_merged_anchor(ws, row, c))
                            for f, c in field_cols.items()})
        band_heights.append(ws.row_dimensions[row].height)

    # Fallback body styles use the first populated data row.
    fallback_row = header_row + 1
    fallback_body = {f: style_snapshot(ws.cell(fallback_row, c))
                     for f, c in field_cols.items()}

    body_by_case = {}
    row_heights = {}
    for case, row in case_rows.items():
        body_by_case[case] = {f: style_snapshot(ws.cell(row, c))
                              for f, c in field_cols.items()}
        row_heights[case] = ws.row_dimensions[row].height

    out = {
        "header": header_styles,
        "fallback_body": fallback_body,
        "body_by_case": body_by_case,
        "bands": band_styles,
        "widths": widths,
        "band_heights": band_heights,
        "header_height": ws.row_dimensions[header_row].height,
        "row_heights": row_heights,
    }
    wb.close()
    return out


def write_band_row(ws, row_idx, values, headers, style_map, ncols):
    """Write a band row, copying template styles and merging contiguous labels."""
    for j, (h, v) in enumerate(zip(headers, values), start=1):
        c = ws.cell(row=row_idx, column=j, value=v)
        if h in style_map:
            apply_style(c, style_map[h])
        else:
            c.font = Font(name=ARIAL, size=9, bold=True)
            c.fill = BAND_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    j = 1
    while j <= ncols:
        k = j
        while (k + 1 <= ncols and values[k] == values[j - 1]
               and values[j - 1] is not None):
            k += 1
        if k > j:
            ws.merge_cells(start_row=row_idx, start_column=j, end_row=row_idx, end_column=k)
        j = k + 1


def resolve_cols(df):
    cid = next((c for c in df.columns if c.lower() == CASE_ID_CANON), None)
    src = next((c for c in df.columns if any(h in c.lower() for h in SOURCE_HINTS)), None)
    return cid, src


def main():
    ap = argparse.ArgumentParser(
        description="Draw the AIFTAX validation audit sample and emit the coding workbook.",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--adjudicated", default=DEFAULT_ADJ,
                    help=f"adjudicated dataset workbook (default: {DEFAULT_ADJ})")
    ap.add_argument("--annotator-a", default=DEFAULT_A,
                    help="pre-adjudication workbook, model A")
    ap.add_argument("--annotator-b", default=DEFAULT_B,
                    help="pre-adjudication workbook, model B")
    ap.add_argument("--protocol-file", default="AIFTAX_Human_Audit_Protocol_2026-08-27.md",
                    help="pre-draw protocol file; its SHA-256 is recorded when found")
    ap.add_argument("--output-dir", default=DEFAULT_OUT,
                    help=f"directory for the draw outputs (default: {DEFAULT_OUT})")
    ap.add_argument("--project-root", default=None,
                    help="override project-root discovery")
    ap.add_argument("--sheet", default=None, help="explicit sheet name in the adjudicated workbook")
    ap.add_argument("--n-random", type=int, default=N_RANDOM, help=f"random draw size (default: {N_RANDOM})")
    ap.add_argument("--seed", type=int, default=SAMPLE_SEED, help=f"draw seed (default: {SAMPLE_SEED})")
    ap.add_argument("--order-seed", type=int, default=ORDER_SEED,
                    help=f"coding-sheet row-order seed (default: {ORDER_SEED})")
    ap.add_argument("--dry-run", action="store_true", help="inspect and report; write nothing")
    args = ap.parse_args()

    root = find_project_root(args.project_root)
    adj_path = resolve_input(args.adjudicated, root, "--adjudicated")
    a_path = resolve_input(args.annotator_a, root, "--annotator-a")
    b_path = resolve_input(args.annotator_b, root, "--annotator-b")
    # Annotator A is the sole style source. Only formatting is copied; no annotation
    # values are transferred into the blinded audit workbook.
    style_path = a_path
    protocol_path = None
    if args.protocol_file:
        p = Path(args.protocol_file).expanduser()
        candidates = [p] if p.is_absolute() else [Path.cwd() / p, root / p]
        protocol_path = next((x.resolve() for x in dict.fromkeys(candidates) if x.exists()), None)
    out_dir = resolve_output(args.output_dir, root)

    def rel(p):
        try:
            return p.relative_to(root)
        except ValueError:
            return p

    print(f"Project root : {root}")
    print(f"Adjudicated  : {rel(adj_path)}")
    print(f"Annotator A  : {rel(a_path)}")
    print(f"Annotator B  : {rel(b_path)}")
    print(f"Style source : {rel(style_path)}")
    if protocol_path:
        print(f"Protocol     : {rel(protocol_path)}  SHA256={sha256_file(protocol_path)[:16]}...")
    else:
        print("Protocol     : NOT FOUND -- confirm the Markdown protocol was committed before drawing")
    print(f"Output dir   : {rel(out_dir)}\n")

    df, sheet_used, hdr_row = load_sheet(adj_path, args.sheet, announce=True)
    cid, src = resolve_cols(df)
    if cid is None:
        sys.exit(f"No 'Case ID' column on sheet '{sheet_used}'. Found: {list(df.columns)}")
    if src is None:
        sys.exit(f"No source-link column on sheet '{sheet_used}'. Found: {list(df.columns)}")

    substantive = [c for c in df.columns if c not in (cid, src)]
    print(f"\nUsing sheet '{sheet_used}': {len(df)} rows, {len(substantive)} substantive fields.")
    if len(substantive) != EXPECT_SUBSTANTIVE:
        print(f"  WARNING: expected {EXPECT_SUBSTANTIVE} substantive fields, found "
              f"{len(substantive)}. Stray columns would be coded as if they were fields:")
        for c in substantive:
            print(f"      - {c}")

    # Visual template: copy only formatting/group-band presentation from the initial
    # annotation workbook. No prior annotation values are copied into audit fields.
    style_df, style_sheet, style_hdr = load_sheet(style_path)
    bands = read_header_bands(style_path, style_sheet, style_hdr)
    template_styles = read_template_styles(style_path, style_sheet, style_hdr)
    if bands:
        gm = bands[-1]
        ngroups = len(dict.fromkeys(gm.get(f) for f in substantive if gm.get(f)))
        print(f"  Header band: {len(bands)} row(s), {ngroups} analytical groups; "
              f"colors/styles copied from {style_path.name} ('{style_sheet}').")
    else:
        print("  Header band: none found in the Annotator A workbook.")

    # ---- draw -------------------------------------------------------------
    ids = df[cid].tolist()
    if len(set(ids)) != len(ids):
        sys.exit("Duplicate Case IDs on the selected sheet. Fix before drawing.")
    if args.n_random < 1 or args.n_random > len(ids):
        sys.exit(f"--n-random must be between 1 and {len(ids)}.")

    rng = random.Random(args.seed)
    sampled = rng.sample(ids, args.n_random)
    # Preserve corpus order in the frozen sample list; coding order is shuffled separately.
    sample_ids = sorted(sampled, key=ids.index)

    print(f"\nDraw (seed {args.seed}):")
    print(f"  simple random sample without replacement : {len(sample_ids)}")
    print(f"  all substantive fields to code           : {len(substantive)}")
    print(f"  total incident-field cells                : {len(sample_ids) * len(substantive)}")
    print(f"  case ids                                  : {sample_ids}")

    # ---- cell provenance + Table V integrity check ------------------------
    a, sa, _ = load_sheet(a_path)
    b, sb, _ = load_sheet(b_path)
    acid, _ = resolve_cols(a)
    bcid, _ = resolve_cols(b)
    if acid is None or bcid is None:
        sys.exit("No 'Case ID' column in one of the pre-adjudication workbooks.")
    a, b = a.set_index(acid), b.set_index(bcid)
    shared = [f for f in substantive if f in a.columns and f in b.columns]

    print(f"\nPre-adjudication pair: {a_path.name} ('{sa}')  vs  {b_path.name} ('{sb}')")
    print("  Full-sample disagreements vs Table V:")
    mismatch = False
    common = [i for i in ids if i in a.index and i in b.index]
    for var, expected in EXPECTED_DIFFS.items():
        col = next((f for f in shared if f.lower() == var.lower()), None)
        if col is None:
            print(f"    {var:<22} column not found in both workbooks")
            mismatch = True
            continue
        got = sum(1 for i in common
                  if norm_for_compare(col, a.at[i, col]) != norm_for_compare(col, b.at[i, col]))
        flag = "ok" if got == expected else f"MISMATCH (paper reports {expected})"
        print(f"    {var:<22} {got:>3}   {flag}")
        mismatch |= (got != expected)
    if mismatch:
        print("  >> These may not be the unchanged pre-adjudication workbooks, or a column name")
        print("     differs. The consensus-vs-adjudicated split is only meaningful if these")
        print("     counts reproduce Table V. Resolve before freezing the draw.")
    else:
        print("  >> All four reproduce Table V. Provenance flags are trustworthy.")

    provenance = {
        i: {
            f: ("consensus"
                if norm_for_compare(f, a.at[i, f]) == norm_for_compare(f, b.at[i, f])
                else "adjudicated")
            for f in shared
        }
        for i in sample_ids if i in a.index and i in b.index
    }
    print(f"  Provenance flags: {len(provenance)} cases x {len(shared)} shared fields.")

    if args.dry_run:
        print("\nDry run: nothing written. Re-run without --dry-run to freeze the draw.")
        return

    # ---- write -------------------------------------------------------------
    sealed = out_dir / "SEALED"
    out_dir.mkdir(parents=True, exist_ok=True)
    sealed.mkdir(parents=True, exist_ok=True)

    order = list(sample_ids)
    random.Random(args.order_seed).shuffle(order)
    idx = df.set_index(cid)

    wb = Workbook()
    ins = wb.active
    ins.title = "Instructions"
    lines = [
        ("AIFTAX author-coded validation audit - coding workbook", True),
        ("", False),
        ("Performed by an author of the study. Code each case from the linked sources in the", False),
        ("controlled source set, using the frozen coding manual only. Where you recognise a case,", False),
        ("code what the sources support, not what you recall deciding earlier.", False),
        ("", False),
        ("Code all 32 substantive fields on the 'Audit' sheet. Do not add or reorder columns.", False),
        ("Do not open the adjudicated workbook, the model outputs, or the key file.", False),
        ("Work down the sheet in the order given; that order is randomised deliberately.", False),
        ("Use the applicable IE/NA code where the evidence does not support a field. Do not guess.", False),
        ("Follow the manual's cross-field consistency rules (manifestation vs impact, risk vs", False),
        ("recovery, industry vs risk domain, failure mode vs causal location).", False),
        ("", False),
        ("Format example (fictional case, NOT part of the audit):", True),
        ("  Failure Category: Operational  |  Propagation Reach: P2  |  Risk Assessment: High", False),
        ("  Multi-label: semicolon-separated, unordered, e.g. 'Financial/economic; Service/operational'", False),
        ("  Narrative: neutral, source-grounded wording; one or two sentences.", False),
        ("", False),
        (f"Draw seed {args.seed} | order seed {args.order_seed} | "
         f"generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}", False),
    ]
    for r, (text, bold) in enumerate(lines, start=1):
        ins.cell(row=r, column=1, value=text).font = Font(name=ARIAL, size=11, bold=bold)
    ins.column_dimensions["A"].width = 108

    aud = wb.create_sheet("Audit")
    headers = [cid, src] + substantive

    for bi, bmap in enumerate(bands, start=1):
        band_style = (template_styles["bands"][bi - 1]
                      if bi - 1 < len(template_styles["bands"]) else {})
        write_band_row(
            aud, bi, [bmap.get(h) for h in headers], headers, band_style, len(headers)
        )
        if bi - 1 < len(template_styles["band_heights"]):
            height = template_styles["band_heights"][bi - 1]
            if height is not None:
                aud.row_dimensions[bi].height = height

    hrow = len(bands) + 1
    for j, h in enumerate(headers, start=1):
        c = aud.cell(row=hrow, column=j, value=h)
        if h in template_styles["header"]:
            apply_style(c, template_styles["header"][h])
        else:
            c.font = Font(name=ARIAL, size=10, bold=True)
            c.fill = HEADER_FILL
            c.alignment = Alignment(wrap_text=True, vertical="top")
    if template_styles["header_height"] is not None:
        aud.row_dimensions[hrow].height = template_styles["header_height"]

    for r, case in enumerate(order, start=hrow + 1):
        case_style = template_styles["body_by_case"].get(case, {})
        for j, h in enumerate(headers, start=1):
            value = None
            if h == cid:
                value = case
            elif h == src:
                value = str(idx.at[case, src])

            cell = aud.cell(row=r, column=j, value=value)
            style = case_style.get(h, template_styles["fallback_body"].get(h))
            if style is not None:
                apply_style(cell, style)
            else:
                cell.font = Font(name=ARIAL, size=10)
                if j >= 3:
                    cell.fill = FILLIN_FILL

            # Keep source links readable even if the template uses a different alignment.
            if h == src:
                cell.alignment = copy(cell.alignment)
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="top", wrap_text=True,
                    text_rotation=cell.alignment.text_rotation,
                    shrink_to_fit=cell.alignment.shrink_to_fit,
                    indent=cell.alignment.indent
                )

        height = template_styles["row_heights"].get(case)
        if height is not None:
            aud.row_dimensions[r].height = height

    aud.freeze_panes = aud.cell(row=hrow + 1, column=3)
    for j, h in enumerate(headers, start=1):
        width = template_styles["widths"].get(h)
        if width is None:
            width = 10 if h == cid else (46 if h == src else 24)
        aud.column_dimensions[get_column_letter(j)].width = width

    try:
        src_wb = load_workbook(adj_path, read_only=True)
        for nm in src_wb.sheetnames:
            if any(k in nm.lower() for k in ("vocab", "controlled", "lists")):
                vs, ws = src_wb[nm], wb.create_sheet("Vocabulary")
                for row in vs.iter_rows(values_only=True):
                    ws.append(list(row))
                print(f"Copied controlled-vocabulary sheet '{nm}'.")
                break
        else:
            print("No vocabulary sheet found; give the coder the manual's vocabularies separately.")
        src_wb.close()
    except Exception as e:
        print(f"Vocabulary copy skipped: {e}")

    coding = out_dir / "AIFTax_Audit_Coding_Workbook.xlsx"
    wb.save(coding)

    key_rows = []
    for case in sample_ids:
        row = {cid: case}
        for f in substantive:
            row[f] = idx.at[case, f]
            if provenance:
                row[f + " [provenance]"] = provenance.get(case, {}).get(f, "")
        key_rows.append(row)
    key = sealed / "aiftax_audit_KEY.xlsx"
    pd.DataFrame(key_rows).to_excel(key, index=False, sheet_name="Key")

    draw_record = {
        "generated_utc": datetime.now(timezone.utc).isoformat(),
        "project_root": str(root),
        "preregistered_protocol": (str(rel(protocol_path)) if protocol_path else None),
        "preregistered_protocol_sha256": (sha256_file(protocol_path) if protocol_path else None),
        "adjudicated_workbook": str(rel(adj_path)),
        "adjudicated_sheet": sheet_used,
        "annotator_a": str(rel(a_path)),
        "annotator_b": str(rel(b_path)),
        "style_source": str(rel(style_path)),
        "style_source_sheet": style_sheet,
        "table_v_check_passed": not mismatch,
        "sample_rule": "simple random sample without replacement from the 100-case adjudicated corpus",
        "sample_seed": args.seed,
        "order_seed": args.order_seed,
        "n_random": len(sample_ids),
        "random_sample_case_ids": sample_ids,
        "coding_order_case_ids": order,
        "n_substantive_fields": len(substantive),
        "n_incident_field_cells": len(sample_ids) * len(substantive),
        "substantive_fields": substantive,
        "header_band_rows": len(bands),
        "field_groups": ({f: bands[-1].get(f) for f in substantive} if bands else {}),
        "comparison_normalization": {
            "single_label_and_text": "trim/collapse whitespace; case-insensitive",
            "multi_label": "semicolon-separated labels compared as unordered normalized sets",
            "multi_label_fields": sorted(MULTILABEL_FIELDS),
        },
        "decision_rules": {
            "dataset_correction": "NONE. Measurement only. Discrepancies are reported; adjudicated "
                                  "labels are not revised on the basis of this audit.",
            "reporting": "All 32 substantive fields are coded and retained. Body emphasis may focus "
                         "on the four principal variables; full field-level results ship in the package.",
            "trigger": "If human-vs-adjudicated exact agreement on Failure Category < 0.70, the "
                       "RQ1a distribution is reported as provisional and the abstract reframed.",
            "statistics": {
                "nominal": "exact agreement; unweighted Cohen kappa; PABAK where marginals are skewed",
                "ordinal": "exact agreement; unweighted and quadratic-weighted Cohen kappa",
                "multi_label": "per-label Cohen kappa, summarized by macro average; optional exact-set/Jaccard diagnostics may be retained in the package",
                "narrative": "normalised exact agreement as diagnostic only; not interpreted as categorical disagreement",
                "ci": "2,000 nonparametric bootstrap resamples, fixed seed, matching Section IV-C",
            },
            "secondary_outcome": "Human IE/NA rate across all coded cells, compared against the "
                                 "adjudicated dataset's 3/3,200 (0.09%).",
            "provenance_split": "Agreement reported separately for cells where the two models agreed "
                                "pre-adjudication vs cells the authors adjudicated.",
        },
    }
    record_path = out_dir / "aiftax_audit_draw_record.json"
    record_path.write_text(json.dumps(draw_record, indent=2, default=str))

    print(f"\nWrote:\n  {rel(coding)}   <- code in this one"
          f"\n  {rel(record_path)}   <- post-draw reproducibility record"
          f"\n  {rel(key)}   <- SEALED; move it off this machine or hand it to a colleague")


if __name__ == "__main__":
    main()