#!/usr/bin/env python3
"""
Normalize the 34-column AIFTax final adjudicated corpus for analysis and figures.

The final adjudicated XLSX/CSV remains the authoritative dataset. This script
creates separate analysis-ready outputs without overwriting any corpus fields.

Default outputs:
  - clean_data.json
      Compact records expected by figures.py, plus additional derived fields.
  - AIFTax_Analysis_Data.csv
      All 34 original fields with transparent analysis fields appended.
  - normalization_summary.json
      Counts and metadata used to verify paper figures and tables.

Example from an Analysis directory:
    python normalize_data.py \
      --input "../Final Corpus/AIFTax_Final_Adjudicated_Corpus.xlsx"

The script also supports the two-row CSV/XLSX layout in which the first row
contains group headers and the second row contains the 34 field names.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence


EXPECTED_COLUMNS = [
    "Case ID",
    "Source Link(s) / Citation(s)",
    "Date",
    "Entity",
    "System Type",
    "Title",
    "Industry",
    "Application Domain",
    "Failure Category",
    "Failure Mode(s)",
    "Manifestation Pattern",
    "Manifestation — Brief Description",
    "Propagation Reach",
    "Boundary Transfer",
    "Transfer Mode",
    "Amplification",
    "Propagation Chain",
    "Detection Timing",
    "Detector",
    "Detection Signal",
    "Impact Dimension(s)",
    "Risk Assessment",
    "Risk Domain",
    "Risk Subdomain",
    "Recovery Complexity",
    "Recovery Evidence Status",
    "Recovery / Maintenance Action(s)",
    "Causal Location(s)",
    "Causal Mechanism",
    "Missing Safeguard",
    "Root Cause Description",
    "Lifecycle Phase",
    "Temporal Pattern",
    "Timing / Period",
]

HEADER_ALIASES = {
    "Impact  Dimension(s)": "Impact Dimension(s)",
    "Notes / Flag (optinal)": "Notes / Flag (optional)",
}

PRIMARY_CATEGORIES = [
    "Operational",
    "Distributional",
    "Adversarial",
    "Mixed/Hybrid",
]

CATEGORY_ALIASES = {
    "operational": "Operational",
    "distributional": "Distributional",
    "adversarial": "Adversarial",
    "mixed/hybrid": "Mixed/Hybrid",
    "mixed / hybrid": "Mixed/Hybrid",
    "mixed-hybrid": "Mixed/Hybrid",
}

# Detailed analytical group used by Figure 1. In the final schema, the broad
# category and detailed failure mode are separate fields.
MODE_TO_SUBCATEGORY = {
    # Operational
    "Capability Limitation": "Capability Limitation",
    "Design/Implementation Defect": "Design & Implementation",
    "Infrastructure/Dependency Failure": "Design & Implementation",
    "Objective/Metric Misalignment": "Design & Implementation",
    "Safety Validation/Oversight Gap": "Safety Validation",
    "Content Generation/Filtering Failure": "Content Generation",
    "Information Quality/Input Handling Failure": "Information Quality",

    # Distributional
    "Bias & Discrimination": "Bias & Discrimination",
    "Hallucination/Confabulation": "Hallucination",
    "Representation/Low-Resource Gap": "Prediction & Drift",
    "Data/Domain Shift": "Prediction & Drift",

    # Adversarial
    "Security Exploitation": "Security Exploitation",
    "Disinformation/Manipulation": "Disinformation & Manipulation",
    "Nation-State/Military Exploitation": "Disinformation & Manipulation",
    "Identity/Impersonation/Deepfake": "Identity & Targeting",
    "Targeting/Surveillance Abuse": "Identity & Targeting",
}

ALLOWED_MODES_BY_CATEGORY = {
    "Operational": {
        "Capability Limitation",
        "Design/Implementation Defect",
        "Infrastructure/Dependency Failure",
        "Objective/Metric Misalignment",
        "Safety Validation/Oversight Gap",
        "Content Generation/Filtering Failure",
        "Information Quality/Input Handling Failure",
    },
    "Distributional": {
        "Bias & Discrimination",
        "Hallucination/Confabulation",
        "Representation/Low-Resource Gap",
        "Data/Domain Shift",
    },
    "Adversarial": {
        "Security Exploitation",
        "Disinformation/Manipulation",
        "Nation-State/Military Exploitation",
        "Identity/Impersonation/Deepfake",
        "Targeting/Surveillance Abuse",
    },
}

RISK_ORDER = ["Critical", "Severe", "High", "Moderate", "Low"]
RISK_ALIASES = {
    "critical": "Critical",
    "critical risk": "Critical",
    "severe": "Severe",
    "severe risk": "Severe",
    "high": "High",
    "high risk": "High",
    "moderate": "Moderate",
    "moderate risk": "Moderate",
    "low": "Low",
    "low risk": "Low",
}

RECOVERY_ORDER = ["Very High", "Severe", "High", "Moderate", "Low"]
RECOVERY_ALIASES = {
    "very high": "Very High",
    "severe": "Severe",
    "high": "High",
    "moderate to high": "High",
    "moderate-to-high": "High",
    "moderate": "Moderate",
    "low": "Low",
}

INDUSTRY_ORDER = [
    "Healthcare & Medicine",
    "Government & Public Services",
    "Social Media & Platforms",
    "Transportation & Automotive",
    "Technology & AI Services",
    "Security & Law Enforcement",
    "Finance, Retail & Commerce",
    "Education & Workforce",
    "Other",
]

PROPAGATION_MAP = {
    "P1": "Single downstream component/decision",
    "P2": "Multi-component/workflow/system-wide",
    "P3": "Cross-organizational/ecosystem/public-scale",
}

ANALYSIS_COLUMNS = [
    "Incident Year",
    "Analysis Primary Category",
    "Analysis Subcategory",
    "Analysis Risk Level",
    "Analysis Sector",
    "Analysis Recovery Level",
    "Propagation Level",
    "Propagation Label",
]


class NormalizationError(ValueError):
    """Raised when the final corpus violates the expected analysis schema."""


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u00a0", " ")).strip()


def normalize_header(value: Any) -> str:
    text = normalize_space(value)
    return HEADER_ALIASES.get(text, text)


def split_multilabel(value: Any) -> list[str]:
    return [
        normalize_space(part)
        for part in normalize_space(value).split(";")
        if normalize_space(part)
    ]


def normalize_case_id(value: Any) -> str:
    if value is None or normalize_space(value) == "":
        raise NormalizationError("Missing Case ID")
    if isinstance(value, bool):
        raise NormalizationError(f"Invalid Case ID: {value!r}")
    try:
        numeric = float(value)
        if numeric.is_integer():
            return str(int(numeric))
    except (TypeError, ValueError):
        pass
    text = normalize_space(value)
    if not text:
        raise NormalizationError("Missing Case ID")
    return text


def excel_serial_to_date(value: float) -> date:
    # Import only when needed so CSV-only use does not require openpyxl.
    try:
        from openpyxl.utils.datetime import from_excel
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required to interpret Excel serial dates. "
            "Install it with: python -m pip install openpyxl"
        ) from exc
    converted = from_excel(value)
    return converted.date() if isinstance(converted, datetime) else converted


def normalize_date(value: Any) -> str:
    if value is None or normalize_space(value) == "":
        return ""

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, (int, float)) and not isinstance(value, bool):
        # Excel serials for modern incident dates are normally 5-digit values.
        if 20000 <= float(value) <= 80000:
            return excel_serial_to_date(float(value)).isoformat()
        if 1000 <= float(value) <= 9999 and float(value).is_integer():
            return str(int(value))

    text = normalize_space(value)

    # Numeric Excel serial stored as text.
    if re.fullmatch(r"\d{5}(?:\.0+)?", text):
        return excel_serial_to_date(float(text)).isoformat()

    # ISO and common date formats.
    candidates = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m",
        "%Y",
    ]
    for pattern in candidates:
        try:
            parsed = datetime.strptime(text, pattern)
            if pattern == "%Y":
                return f"{parsed.year:04d}"
            if pattern == "%Y-%m":
                return f"{parsed.year:04d}-{parsed.month:02d}"
            return parsed.date().isoformat()
        except ValueError:
            continue

    # Preserve a source-grounded period string but require a visible year.
    if re.search(r"\b(?:19|20)\d{2}\b", text):
        return text

    raise NormalizationError(f"Unrecognized date value: {value!r}")


def extract_year(normalized_date: str) -> str:
    match = re.search(r"\b((?:19|20)\d{2})\b", normalized_date)
    if not match:
        raise NormalizationError(
            f"Could not extract an incident year from date: {normalized_date!r}"
        )
    return match.group(1)


def normalize_category(value: Any) -> str:
    text = normalize_space(value)
    category = CATEGORY_ALIASES.get(text.casefold())
    if category is None:
        raise NormalizationError(f"Unmapped Failure Category: {text!r}")
    return category


def normalize_risk(value: Any) -> str:
    text = normalize_space(value)
    risk = RISK_ALIASES.get(text.casefold())
    if risk is None:
        raise NormalizationError(f"Unmapped Risk Assessment: {text!r}")
    return risk


def normalize_recovery(value: Any) -> str:
    text = normalize_space(value)
    recovery = RECOVERY_ALIASES.get(text.casefold())
    if recovery is None:
        raise NormalizationError(f"Unmapped Recovery Complexity: {text!r}")
    return recovery


def normalize_industry(value: Any) -> str:
    # The final corpus already contains the adjudicated analytical sectors.
    # Do not collapse or replace them with a new heuristic mapping.
    text = normalize_space(value)
    if text not in INDUSTRY_ORDER:
        raise NormalizationError(f"Unmapped Industry: {text!r}")
    return text


def parse_propagation(value: Any) -> tuple[str, str]:
    text = normalize_space(value)
    match = re.match(r"^(P[1-3])\s*[—–-]\s*(.+)$", text)
    if not match:
        raise NormalizationError(f"Unmapped Propagation Reach: {text!r}")
    level, label = match.group(1), normalize_space(match.group(2))
    expected = PROPAGATION_MAP[level]
    if label.casefold() != expected.casefold():
        raise NormalizationError(
            f"Unexpected Propagation Reach label for {level}: {label!r}; "
            f"expected {expected!r}"
        )
    return level, expected


def derive_subcategory(category: str, modes: Sequence[str]) -> str:
    if not modes:
        raise NormalizationError("Failure Mode(s) is blank")

    if category == "Mixed/Hybrid":
        operational = any(
            mode in ALLOWED_MODES_BY_CATEGORY["Operational"] for mode in modes
        )
        distributional = any(
            mode in ALLOWED_MODES_BY_CATEGORY["Distributional"] for mode in modes
        )
        if not (operational and distributional):
            raise NormalizationError(
                "Mixed/Hybrid cases must include at least one Operational and "
                "one Distributional failure mode"
            )
        return "Distributional + Operational"

    allowed = ALLOWED_MODES_BY_CATEGORY[category]
    unexpected = [mode for mode in modes if mode not in allowed]
    if unexpected:
        raise NormalizationError(
            f"Failure mode(s) {unexpected!r} are inconsistent with "
            f"Failure Category {category!r}"
        )
    if len(modes) != 1:
        raise NormalizationError(
            f"{category} cases are expected to have one failure mode; found {modes!r}"
        )

    mode = modes[0]
    try:
        return MODE_TO_SUBCATEGORY[mode]
    except KeyError as exc:
        raise NormalizationError(f"Unmapped Failure Mode: {mode!r}") from exc


def locate_header(rows: Sequence[Sequence[Any]]) -> tuple[int, list[str]]:
    for index, row in enumerate(rows[:10]):
        headers = [normalize_header(value) for value in row]
        if (
            "Case ID" in headers
            and "Failure Category" in headers
            and "Timing / Period" in headers
        ):
            return index, headers
    raise NormalizationError(
        "Could not find the 34-column header row containing Case ID, "
        "Failure Category, and Timing / Period."
    )


def validate_schema(headers: Sequence[str]) -> None:
    missing = [column for column in EXPECTED_COLUMNS if column not in headers]
    if missing:
        raise NormalizationError(
            "Missing required final-corpus columns: " + ", ".join(missing)
        )

    actual_order = [header for header in headers if header in EXPECTED_COLUMNS]
    if actual_order != EXPECTED_COLUMNS:
        raise NormalizationError(
            "The final-corpus columns are present but not in the expected "
            "34-column order."
        )


def matrix_to_records(matrix: Sequence[Sequence[Any]]) -> list[dict[str, Any]]:
    header_index, headers = locate_header(matrix)
    validate_schema(headers)

    records: list[dict[str, Any]] = []
    for row in matrix[header_index + 1 :]:
        padded = list(row) + [None] * max(0, len(headers) - len(row))
        record = dict(zip(headers, padded[: len(headers)]))
        if normalize_space(record.get("Case ID")):
            records.append(record)
    return records


def load_xlsx(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError(
            "openpyxl is required for XLSX input. Install it with: "
            "python -m pip install openpyxl"
        ) from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise NormalizationError(
            f"Worksheet {sheet_name!r} was not found. Available worksheets: "
            + ", ".join(workbook.sheetnames)
        )
    worksheet = workbook[sheet_name]
    matrix = [list(row) for row in worksheet.iter_rows(values_only=True)]
    return matrix_to_records(matrix)


def load_csv_file(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        matrix = list(csv.reader(handle))
    return matrix_to_records(matrix)


def load_records(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    suffix = path.suffix.casefold()
    if suffix == ".xlsx":
        return load_xlsx(path, sheet_name)
    if suffix == ".csv":
        return load_csv_file(path)
    raise NormalizationError(
        f"Unsupported input format {path.suffix!r}; use .xlsx or .csv"
    )


def discover_default_input(script_dir: Path) -> Path:
    candidates = [
        script_dir / "AIFTax_Final_Adjudicated_Dataset.xlsx",
        script_dir / "AIFTax_Final_Adjudicated_Corpus.xlsx",
        script_dir / "AIFTax_Final_Adjudicated_Corpus.csv",
        script_dir.parent / "Final Corpus" / "AIFTax_Final_Adjudicated_Corpus.xlsx",
        script_dir.parent / "Final Corpus" / "AIFTax_Final_Adjudicated_Corpus.csv",
        Path.cwd() / "AIFTax_Final_Adjudicated_Dataset.xlsx",
        Path.cwd() / "Final Corpus" / "AIFTax_Final_Adjudicated_Corpus.xlsx",
        Path.cwd() / "Final Corpus" / "AIFTax_Final_Adjudicated_Corpus.csv",
    ]
    for candidate in candidates:
        if candidate.exists():
            return candidate.resolve()
    raise FileNotFoundError(
        "No final corpus was found automatically. Supply one with "
        "--input PATH_TO_FINAL_CORPUS.xlsx"
    )


def build_analysis_records(
    source_records: Iterable[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    clean_json: list[dict[str, Any]] = []
    analysis_csv: list[dict[str, str]] = []
    seen_ids: set[str] = set()
    errors: list[str] = []

    for row_number, source in enumerate(source_records, start=1):
        try:
            case_id = normalize_case_id(source.get("Case ID"))
            if case_id in seen_ids:
                raise NormalizationError(f"Duplicate Case ID: {case_id}")
            seen_ids.add(case_id)

            normalized_date = normalize_date(source.get("Date"))
            year = extract_year(normalized_date)
            category = normalize_category(source.get("Failure Category"))
            modes = split_multilabel(source.get("Failure Mode(s)"))
            subcategory = derive_subcategory(category, modes)
            risk = normalize_risk(source.get("Risk Assessment"))
            sector = normalize_industry(source.get("Industry"))
            recovery = normalize_recovery(source.get("Recovery Complexity"))
            propagation_level, propagation_label = parse_propagation(
                source.get("Propagation Reach")
            )

            original = {
                column: (
                    normalized_date
                    if column == "Date"
                    else normalize_space(source.get(column))
                )
                for column in EXPECTED_COLUMNS
            }
            original["Case ID"] = case_id

            clean_json.append(
                {
                    # Backward-compatible keys used by figures.py
                    "id": case_id,
                    "year": year,
                    "entity": original["Entity"],
                    "title": original["Title"],
                    "raw_category": original["Failure Category"],
                    "primary_category": category,
                    "subcategory": subcategory,
                    "risk": risk,
                    "raw_industry": original["Industry"],
                    "sector": sector,
                    "recovery": recovery,
                    "propagation": original["Propagation Reach"],
                    "failure_mode": original["Failure Mode(s)"],

                    # Additional transparent fields for revised analyses
                    "case_id": case_id,
                    "source_link": original["Source Link(s) / Citation(s)"],
                    "date": normalized_date,
                    "system_type": original["System Type"],
                    "failure_modes": modes,
                    "propagation_level": propagation_level,
                    "propagation_label": propagation_label,
                    "temporal_pattern": original["Temporal Pattern"],
                }
            )

            csv_record = dict(original)
            csv_record.update(
                {
                    "Incident Year": year,
                    "Analysis Primary Category": category,
                    "Analysis Subcategory": subcategory,
                    "Analysis Risk Level": risk,
                    "Analysis Sector": sector,
                    "Analysis Recovery Level": recovery,
                    "Propagation Level": propagation_level,
                    "Propagation Label": propagation_label,
                }
            )
            analysis_csv.append(csv_record)

        except (NormalizationError, ValueError, TypeError) as exc:
            case_hint = normalize_space(source.get("Case ID")) or f"row {row_number}"
            errors.append(f"Case {case_hint}: {exc}")

    if errors:
        message = "Normalization failed:\n  - " + "\n  - ".join(errors)
        raise NormalizationError(message)

    return clean_json, analysis_csv


def ordered_counts(values: Iterable[str], order: Sequence[str]) -> dict[str, int]:
    counts = Counter(values)
    return {label: counts.get(label, 0) for label in order}


def make_summary(
    input_path: Path,
    clean: Sequence[dict[str, Any]],
) -> dict[str, Any]:
    years = Counter(row["year"] for row in clean)
    subcategories = Counter(row["subcategory"] for row in clean)
    sectors = Counter(row["sector"] for row in clean)
    propagation = Counter(row["propagation_level"] for row in clean)
    system_types = Counter(row["system_type"] for row in clean)
    temporal = Counter(row["temporal_pattern"] for row in clean)

    return {
        "input_file": str(input_path),
        "cases": len(clean),
        "primary_categories": ordered_counts(
            (row["primary_category"] for row in clean), PRIMARY_CATEGORIES
        ),
        "subcategories": dict(subcategories.most_common()),
        "risk_levels": ordered_counts(
            (row["risk"] for row in clean), RISK_ORDER
        ),
        "sectors": {
            label: sectors.get(label, 0)
            for label in INDUSTRY_ORDER
            if sectors.get(label, 0)
        },
        "recovery_levels": ordered_counts(
            (row["recovery"] for row in clean), RECOVERY_ORDER
        ),
        "propagation_levels": {
            level: propagation.get(level, 0)
            for level in ["P1", "P2", "P3"]
        },
        "system_types": dict(system_types.most_common()),
        "temporal_patterns": dict(temporal.most_common()),
        "years": dict(sorted(years.items())),
    }


def print_summary(summary: dict[str, Any]) -> None:
    def section(title: str) -> None:
        print(f"\n{'─' * 64}\n{title}\n{'─' * 64}")

    total = summary["cases"]
    section("PRIMARY CATEGORY COUNTS")
    for label, count in summary["primary_categories"].items():
        print(f"  {label:<22} {count:>3}  ({count / total * 100:.0f}%)")

    section("SUBCATEGORY COUNTS")
    for label, count in summary["subcategories"].items():
        print(f"  {label:<40} {count:>3}")

    section("RISK SEVERITY COUNTS")
    for label, count in summary["risk_levels"].items():
        print(f"  {label:<12} {count:>3}  ({count / total * 100:.0f}%)")
    high_or_worse = sum(
        summary["risk_levels"][label]
        for label in ["Critical", "Severe", "High"]
    )
    print(
        f"  {'High-or-worse':<12} {high_or_worse:>3}  "
        f"({high_or_worse / total * 100:.0f}%)"
    )

    section("SECTOR COUNTS")
    for label, count in summary["sectors"].items():
        print(f"  {label:<35} {count:>3}  ({count / total * 100:.0f}%)")

    section("RECOVERY COMPLEXITY COUNTS")
    for label, count in summary["recovery_levels"].items():
        print(f"  {label:<12} {count:>3}  ({count / total * 100:.0f}%)")
    high_or_worse_recovery = sum(
        summary["recovery_levels"][label]
        for label in ["Very High", "Severe", "High"]
    )
    print(
        f"  {'High-or-worse':<12} {high_or_worse_recovery:>3}  "
        f"({high_or_worse_recovery / total * 100:.0f}%)"
    )

    section("PROPAGATION REACH")
    for level, count in summary["propagation_levels"].items():
        print(f"  {level:<4} {count:>3}  ({count / total * 100:.0f}%)")

    section("YEAR DISTRIBUTION")
    for year, count in summary["years"].items():
        print(f"  {year}  {count:>3}  {'█' * count}")


def write_outputs(
    output_dir: Path,
    clean: Sequence[dict[str, Any]],
    analysis_rows: Sequence[dict[str, str]],
    summary: dict[str, Any],
    json_name: str,
    csv_name: str,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)

    clean_path = output_dir / json_name
    with clean_path.open("w", encoding="utf-8") as handle:
        json.dump(clean, handle, indent=2, ensure_ascii=False)

    analysis_path = output_dir / csv_name
    fieldnames = EXPECTED_COLUMNS + ANALYSIS_COLUMNS
    with analysis_path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(analysis_rows)

    summary_path = output_dir / "normalization_summary.json"
    with summary_path.open("w", encoding="utf-8") as handle:
        json.dump(summary, handle, indent=2, ensure_ascii=False)

    print(f"\n✓ clean_data.json saved → {clean_path}")
    print(f"✓ analysis-ready CSV saved → {analysis_path}")
    print(f"✓ normalization summary saved → {summary_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Normalize the 34-column AIFTax final adjudicated corpus into "
            "analysis-ready CSV and JSON outputs."
        )
    )
    parser.add_argument(
        "--input",
        type=Path,
        help=(
            "Path to AIFTax_Final_Adjudicated_Corpus.xlsx or .csv. "
            "When omitted, the script searches common replication-package paths."
        ),
    )
    parser.add_argument(
        "--sheet",
        default="Final Corpus",
        help="XLSX worksheet containing the 34-column corpus (default: Final Corpus).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        help="Output directory (default: directory containing this script).",
    )
    parser.add_argument(
        "--json-name",
        default="clean_data.json",
        help="JSON output filename expected by figures.py.",
    )
    parser.add_argument(
        "--csv-name",
        default="AIFTax_Analysis_Data.csv",
        help="Analysis-ready CSV output filename.",
    )
    parser.add_argument(
        "--expected-cases",
        type=int,
        default=100,
        help="Expected number of cases; use 0 to disable this check (default: 100).",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    script_dir = Path(__file__).resolve().parent

    try:
        input_path = (
            args.input.expanduser().resolve()
            if args.input is not None
            else discover_default_input(script_dir)
        )
        if not input_path.exists():
            raise FileNotFoundError(f"Input file does not exist: {input_path}")

        output_dir = (
            args.output_dir.expanduser().resolve()
            if args.output_dir is not None
            else script_dir
        )

        source_records = load_records(input_path, args.sheet)
        clean, analysis_rows = build_analysis_records(source_records)

        if args.expected_cases and len(clean) != args.expected_cases:
            raise NormalizationError(
                f"Expected {args.expected_cases} cases, found {len(clean)}"
            )

        summary = make_summary(input_path, clean)
        print(f"✓ Loaded and validated {len(clean)} cases from {input_path}")
        print_summary(summary)
        write_outputs(
            output_dir,
            clean,
            analysis_rows,
            summary,
            args.json_name,
            args.csv_name,
        )
        return 0

    except (FileNotFoundError, NormalizationError, RuntimeError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
