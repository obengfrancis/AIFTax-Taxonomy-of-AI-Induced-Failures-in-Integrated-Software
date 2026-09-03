#!/usr/bin/env python3
"""AIFTax inter-rater agreement calculator.

Compares two independent AIFTax annotation workbooks by Case ID and produces:
  * an Excel agreement report;
  * a JSON summary;
  * CSV summaries and disagreement files.

The script treats the following as the principal analytical variables:
  1. Failure Category (nominal Cohen's kappa)
  2. Propagation Reach (quadratic weighted Cohen's kappa)
  3. Risk Assessment (quadratic weighted Cohen's kappa)
  4. Recovery Complexity (quadratic weighted Cohen's kappa)


Dependencies:
  Python 3.10+
  openpyxl 3.1+ 
  
  openpyxl is required. Install it with: python -m pip install openpyxl

Usage/Run:
  python Inter_Rater_Analysis/aiftax_inter_rater_agreement.py \
      --annotator-a "Annotation Materials/AIFTax_Annotator_A_Annotation_Completed.xlsx" \
      --annotator-b "Annotation Materials/AIFTax_Annotator_B_Annotation_Completed.xlsx" \
      --output-dir "Inter_Rater_Analysis/irr_results"
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import random
import re
import statistics
import sys
import unicodedata
from collections import Counter, OrderedDict
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required. Install it with: python -m pip install openpyxl"
    ) from exc


# ---------------------------------------------------------------------------
# Study configuration
# ---------------------------------------------------------------------------

PRINCIPAL_FIELDS: "OrderedDict[str, dict[str, Any]]" = OrderedDict(
    [
        (
            "Failure Category",
            {
                "kind": "nominal",
                "order": [
                    "Operational",
                    "Distributional",
                    "Adversarial",
                    "Mixed/Hybrid",
                    "IE",
                ],
                "primary_metric": "Cohen's kappa",
            },
        ),
        (
            "Propagation Reach",
            {
                "kind": "ordinal",
                "order": ["P0", "P1", "P2", "P3"],
                "primary_metric": "Quadratic weighted kappa",
            },
        ),
        (
            "Risk Assessment",
            {
                "kind": "ordinal",
                "order": ["Low", "Moderate", "High", "Severe", "Critical"],
                "primary_metric": "Quadratic weighted kappa",
            },
        ),
        (
            "Recovery Complexity",
            {
                "kind": "ordinal",
                "order": ["Low", "Moderate", "High", "Severe", "Very High"],
                "primary_metric": "Quadratic weighted kappa",
            },
        ),
    ]
)

SUPPORTING_NOMINAL_FIELDS = [
    "System Type",
    "Industry",
    "Manifestation Pattern",
    "Boundary Transfer",
    "Transfer Mode",
    "Amplification",
    "Detector",
    "Detection Signal",
    "Risk Domain",
    "Recovery Evidence Status",
    "Temporal Pattern",
]

SUPPORTING_ORDINAL_FIELDS: "OrderedDict[str, list[str]]" = OrderedDict(
    [
        ("Detection Timing", ["Immediate", "Delayed", "Retrospective"]),
        ("Lifecycle Phase", ["Pre-deployment", "Deployment/Rollout", "Post-deployment"]),
    ]
)

MULTILABEL_FIELDS = [
    "Failure Mode(s)",
    "Impact Dimension(s)",
    "Recovery / Maintenance Action(s)",
    "Causal Location(s)",
]

TEXT_DIAGNOSTIC_FIELDS = [
    "Date",
    "Entity",
    "Title",
    "Application Domain",
    "Manifestation — Brief Description",
    "Propagation Chain",
    "Risk Subdomain",
    "Causal Mechanism",
    "Missing Safeguard",
    "Root Cause Description",
    "Timing / Period",
]

ADMINISTRATIVE_FIELDS = ["Case ID", "Source Link(s) / Citation(s)"]
OPTIONAL_NONANALYTIC_FIELDS = ["Notes / Flag (optional)"]

HEADER_ALIASES = {
    "case id": "Case ID",
    "source link(s) / citation(s)": "Source Link(s) / Citation(s)",
    "date": "Date",
    "entity": "Entity",
    "system type": "System Type",
    "title": "Title",
    "industry": "Industry",
    "application domain": "Application Domain",
    "failure category": "Failure Category",
    "failure mode(s)": "Failure Mode(s)",
    "failure mode": "Failure Mode(s)",
    "manifestation pattern": "Manifestation Pattern",
    "manifestation — brief description": "Manifestation — Brief Description",
    "manifestation - brief description": "Manifestation — Brief Description",
    "propagation reach": "Propagation Reach",
    "boundary transfer": "Boundary Transfer",
    "transfer mode": "Transfer Mode",
    "amplification": "Amplification",
    "propagation chain": "Propagation Chain",
    "detection timing": "Detection Timing",
    "detector": "Detector",
    "detection signal": "Detection Signal",
    "impact dimension(s)": "Impact Dimension(s)",
    "impact class": "Impact Dimension(s)",
    "risk assessment": "Risk Assessment",
    "risk domain": "Risk Domain",
    "risk subdomain": "Risk Subdomain",
    "recovery complexity": "Recovery Complexity",
    "recovery evidence status": "Recovery Evidence Status",
    "recovery / maintenance action(s)": "Recovery / Maintenance Action(s)",
    "recovery / maintenance actions": "Recovery / Maintenance Action(s)",
    "causal location(s)": "Causal Location(s)",
    "causal location": "Causal Location(s)",
    "causal mechanism": "Causal Mechanism",
    "missing safeguard": "Missing Safeguard",
    "root cause description": "Root Cause Description",
    "lifecycle phase": "Lifecycle Phase",
    "temporal pattern": "Temporal Pattern",
    "timing / period": "Timing / Period",
    "notes / flag (optional)": "Notes / Flag (optional)",
    "notes / flag (optinal)": "Notes / Flag (optional)",
}

LIST_HEADER_TO_FIELD = {
    "System Type": "System Type",
    "Industry": "Industry",
    "Failure Category": "Failure Category",
    "Failure Mode": "Failure Mode(s)",
    "Manifestation Pattern": "Manifestation Pattern",
    "Propagation Reach": "Propagation Reach",
    "Transfer Mode": "Transfer Mode",
    "Detection Timing": "Detection Timing",
    "Detector": "Detector",
    "Detection Signal": "Detection Signal",
    "Impact Class": "Impact Dimension(s)",
    "Risk Assessment": "Risk Assessment",
    "Risk Domain": "Risk Domain",
    "Recovery Complexity": "Recovery Complexity",
    "Recovery Evidence Status": "Recovery Evidence Status",
    "Causal Location": "Causal Location(s)",
    "Lifecycle Phase": "Lifecycle Phase",
    "Temporal Pattern": "Temporal Pattern",
    "Recovery / Maintenance Actions": "Recovery / Maintenance Action(s)",
}

SPECIAL_MISSING_CODES = {"ie", "na", "other", "other / ie"}


@dataclass
class WorkbookData:
    path: Path
    sheet_name: str
    header_row: int
    headers: list[str]
    records: dict[str, dict[str, Any]]
    original_records: dict[str, dict[str, Any]]
    duplicate_case_ids: list[str]
    vocabularies: dict[str, set[str]]


# ---------------------------------------------------------------------------
# Normalization and workbook reading
# ---------------------------------------------------------------------------


def collapse_spaces(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def normalize_unicode(value: str) -> str:
    value = unicodedata.normalize("NFKC", value)
    replacements = {
        "\u00a0": " ",
        "’": "'",
        "‘": "'",
        "“": '"',
        "”": '"',
        "–": "-",
        "—": "-",
        "−": "-",
    }
    for source, target in replacements.items():
        value = value.replace(source, target)
    return value


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    text = collapse_spaces(normalize_unicode(str(value)))
    key = text.casefold()
    return HEADER_ALIASES.get(key, text)


def scalar_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time().replace(microsecond=0) == datetime.min.time():
            return value.date().isoformat()
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return collapse_spaces(normalize_unicode(str(value)))


def canonical_case_id(value: Any) -> str:
    text = scalar_to_text(value)
    if not text:
        return ""
    try:
        number = float(text)
        if number.is_integer():
            return str(int(number))
    except ValueError:
        pass
    return text


def canonical_key(field: str, value: Any) -> str:
    """Canonical comparison key for a scalar categorical field."""
    text = scalar_to_text(value)
    if not text:
        return ""

    if field == "Propagation Reach":
        match = re.match(r"^(P[0-3])\b", text, flags=re.IGNORECASE)
        if match:
            return match.group(1).upper()
        if text.casefold() == "ie":
            return "IE"

    if field in {"Risk Assessment", "Recovery Complexity"}:
        mapping = {
            "low": "Low",
            "moderate": "Moderate",
            "high": "High",
            "severe": "Severe",
            "critical": "Critical",
            "very high": "Very High",
            "ie": "IE",
            "na": "NA",
            "other": "Other",
            "other / ie": "Other / IE",
        }
        return mapping.get(text.casefold(), text)

    if field == "Failure Category":
        mapping = {
            "operational": "Operational",
            "distributional": "Distributional",
            "adversarial": "Adversarial",
            "mixed/hybrid": "Mixed/Hybrid",
            "mixed / hybrid": "Mixed/Hybrid",
            "ie": "IE",
        }
        return mapping.get(text.casefold(), text)

    if field == "Detection Timing":
        mapping = {
            "immediate": "Immediate",
            "delayed": "Delayed",
            "retrospective": "Retrospective",
            "ie": "IE",
        }
        return mapping.get(text.casefold(), text)

    if field == "Lifecycle Phase":
        mapping = {
            "pre-deployment": "Pre-deployment",
            "deployment/rollout": "Deployment/Rollout",
            "deployment / rollout": "Deployment/Rollout",
            "post-deployment": "Post-deployment",
            "ie": "IE",
        }
        return mapping.get(text.casefold(), text)

    # Controlled nominal labels are compared case-insensitively after
    # whitespace and Unicode normalization.
    return text.casefold()


def normalized_text_key(value: Any) -> str:
    text = scalar_to_text(value)
    return text.casefold()


def parse_multilabel(value: Any) -> frozenset[str]:
    text = scalar_to_text(value)
    if not text:
        return frozenset()
    labels = [collapse_spaces(part) for part in text.split(";")]
    return frozenset(label.casefold() for label in labels if label)


def display_multilabel(value: Any) -> str:
    return scalar_to_text(value)


def find_header_row(ws: Any, max_rows: int = 20) -> int:
    for row_index, row in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_rows), values_only=True),
        start=1,
    ):
        values = [normalize_header(value) for value in row]
        if "Case ID" in values and "Source Link(s) / Citation(s)" in values:
            return row_index
    raise ValueError(
        f"Could not locate the annotation header row in sheet {ws.title!r}. "
        "Expected Case ID and Source Link(s) / Citation(s)."
    )


def read_vocabularies(workbook: Any) -> dict[str, set[str]]:
    if "Lists" not in workbook.sheetnames:
        return {}
    ws = workbook["Lists"]
    # Read sequentially because random cell access is very slow in read-only mode.
    # The controlled lists are short; rows beyond 250 are formatting-only in the
    # current workbooks.
    rows = list(
        ws.iter_rows(
            min_row=1,
            max_row=min(ws.max_row, 250),
            max_col=min(ws.max_column, 60),
            values_only=True,
        )
    )
    if not rows:
        return {}
    headers = [
        collapse_spaces(normalize_unicode(str(value))) if value is not None else ""
        for value in rows[0]
    ]
    vocabularies: dict[str, set[str]] = {}
    for column_index, header in enumerate(headers):
        field = LIST_HEADER_TO_FIELD.get(header)
        if not field:
            continue
        values = {
            scalar_to_text(row[column_index])
            for row in rows[1:]
            if column_index < len(row) and scalar_to_text(row[column_index])
        }
        vocabularies[field] = values
    return vocabularies


def read_annotation_workbook(path: Path, sheet_name: str) -> WorkbookData:
    workbook = load_workbook(path, data_only=True, read_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(
            f"Sheet {sheet_name!r} is missing from {path.name}. "
            f"Available sheets: {', '.join(workbook.sheetnames)}"
        )

    ws = workbook[sheet_name]
    header_row = find_header_row(ws)
    headers = [normalize_header(cell.value) for cell in ws[header_row]]

    if len(headers) != len(set(header for header in headers if header)):
        duplicates = [
            header
            for header, count in Counter(header for header in headers if header).items()
            if count > 1
        ]
        raise ValueError(f"Duplicate normalized headers in {path.name}: {duplicates}")

    records: dict[str, dict[str, Any]] = {}
    original_records: dict[str, dict[str, Any]] = {}
    duplicate_ids: list[str] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_map = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        case_id = canonical_case_id(row_map.get("Case ID"))
        if not case_id:
            continue
        if case_id in records:
            duplicate_ids.append(case_id)
            continue
        original_records[case_id] = dict(row_map)
        records[case_id] = dict(row_map)

    vocabularies = read_vocabularies(workbook)
    workbook.close()

    return WorkbookData(
        path=path,
        sheet_name=sheet_name,
        header_row=header_row,
        headers=headers,
        records=records,
        original_records=original_records,
        duplicate_case_ids=duplicate_ids,
        vocabularies=vocabularies,
    )


# ---------------------------------------------------------------------------
# Agreement metrics
# ---------------------------------------------------------------------------


def safe_divide(numerator: float, denominator: float) -> float | None:
    if denominator == 0:
        return None
    return numerator / denominator


def exact_agreement(keys_a: Sequence[Any], keys_b: Sequence[Any]) -> float | None:
    if not keys_a:
        return None
    return sum(a == b for a, b in zip(keys_a, keys_b)) / len(keys_a)


def cohen_kappa(keys_a: Sequence[str], keys_b: Sequence[str]) -> float | None:
    if len(keys_a) != len(keys_b):
        raise ValueError("Rater vectors must have equal length")
    n = len(keys_a)
    if n == 0:
        return None

    observed = sum(a == b for a, b in zip(keys_a, keys_b)) / n
    counts_a = Counter(keys_a)
    counts_b = Counter(keys_b)
    labels = set(counts_a) | set(counts_b)
    expected = sum((counts_a[label] / n) * (counts_b[label] / n) for label in labels)
    denominator = 1.0 - expected
    if math.isclose(denominator, 0.0, abs_tol=1e-12):
        return None  # 1.0 if math.isclose(observed, 1.0, abs_tol=1e-12) else None
    return (observed - expected) / denominator


def weighted_kappa(
    keys_a: Sequence[str],
    keys_b: Sequence[str],
    order: Sequence[str],
    weighting: str = "quadratic",
) -> float | None:
    if len(keys_a) != len(keys_b):
        raise ValueError("Rater vectors must have equal length")
    n = len(keys_a)
    k = len(order)
    if n == 0 or k < 2:
        return None

    index = {label: position for position, label in enumerate(order)}
    matrix = [[0.0 for _ in range(k)] for _ in range(k)]
    for a, b in zip(keys_a, keys_b):
        if a not in index or b not in index:
            raise ValueError("Weighted kappa received a label outside the supplied order")
        matrix[index[a]][index[b]] += 1.0

    row_totals = [sum(row) for row in matrix]
    col_totals = [sum(matrix[row][col] for row in range(k)) for col in range(k)]

    observed_disagreement = 0.0
    expected_disagreement = 0.0
    for i in range(k):
        for j in range(k):
            distance = abs(i - j) / (k - 1)
            if weighting == "linear":
                weight = distance
            elif weighting == "quadratic":
                weight = distance**2
            else:
                raise ValueError(f"Unknown weighting: {weighting}")
            observed_disagreement += weight * (matrix[i][j] / n)
            expected_disagreement += weight * ((row_totals[i] * col_totals[j]) / (n * n))

    if math.isclose(expected_disagreement, 0.0, abs_tol=1e-12):
        return None #1.0 if math.isclose(observed_disagreement, 0.0, abs_tol=1e-12) else None
    return 1.0 - (observed_disagreement / expected_disagreement)


def percentile(values: Sequence[float], probability: float) -> float | None:
    if not values:
        return None
    ordered = sorted(values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * probability
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] * (1 - fraction) + ordered[upper] * fraction


def bootstrap_ci(
    pairs: Sequence[tuple[str, str]],
    metric: Callable[[Sequence[str], Sequence[str]], float | None],
    iterations: int,
    seed: int,
) -> tuple[float | None, float | None, int]:
    if iterations <= 0 or not pairs:
        return None, None, 0
    rng = random.Random(seed)
    estimates: list[float] = []
    n = len(pairs)
    for _ in range(iterations):
        sample = [pairs[rng.randrange(n)] for _ in range(n)]
        estimate = metric([item[0] for item in sample], [item[1] for item in sample])
        if estimate is not None and math.isfinite(estimate):
            estimates.append(estimate)
    return percentile(estimates, 0.025), percentile(estimates, 0.975), len(estimates)


def confusion_matrix(
    keys_a: Sequence[str], keys_b: Sequence[str], preferred_order: Sequence[str] | None = None
) -> tuple[list[str], list[list[int]]]:
    labels = list(preferred_order or [])
    extras = sorted((set(keys_a) | set(keys_b)) - set(labels))
    labels.extend(extras)
    index = {label: position for position, label in enumerate(labels)}
    matrix = [[0 for _ in labels] for _ in labels]
    for a, b in zip(keys_a, keys_b):
        matrix[index[a]][index[b]] += 1
    return labels, matrix


def multilabel_metrics(
    sets_a: Sequence[frozenset[str]], sets_b: Sequence[frozenset[str]]
) -> dict[str, Any]:
    n = len(sets_a)
    if n == 0:
        return {
            "n": 0,
            "exact": None,
            "mean_jaccard": None,
            "micro_f1": None,
            "macro_label_kappa": None,
            "label_count": 0,
            "per_label": [],
        }

    exact = sum(a == b for a, b in zip(sets_a, sets_b)) / n
    jaccards: list[float] = []
    intersection_total = 0
    size_a_total = 0
    size_b_total = 0
    labels = sorted(set().union(*sets_a, *sets_b))

    for a, b in zip(sets_a, sets_b):
        union = a | b
        jaccards.append(1.0 if not union else len(a & b) / len(union))
        intersection_total += len(a & b)
        size_a_total += len(a)
        size_b_total += len(b)

    denominator = size_a_total + size_b_total
    micro_f1 = 1.0 if denominator == 0 else (2 * intersection_total) / denominator

    per_label: list[dict[str, Any]] = []
    valid_kappas: list[float] = []
    for label in labels:
        vector_a = ["1" if label in value else "0" for value in sets_a]
        vector_b = ["1" if label in value else "0" for value in sets_b]
        kappa = cohen_kappa(vector_a, vector_b)
        if kappa is not None:
            valid_kappas.append(kappa)
        per_label.append(
            {
                "label": label,
                "a_positive": sum(value == "1" for value in vector_a),
                "b_positive": sum(value == "1" for value in vector_b),
                "binary_agreement": exact_agreement(vector_a, vector_b),
                "kappa": kappa,
            }
        )

    return {
        "n": n,
        "exact": exact,
        "mean_jaccard": statistics.fmean(jaccards),
        "micro_f1": micro_f1,
        "macro_label_kappa": statistics.fmean(valid_kappas) if valid_kappas else None,
        "label_count": len(labels),
        "per_label": per_label,
    }


# ---------------------------------------------------------------------------
# Analysis assembly
# ---------------------------------------------------------------------------


def paired_case_ids(data_a: WorkbookData, data_b: WorkbookData) -> list[str]:
    common = set(data_a.records) & set(data_b.records)

    def sort_key(case_id: str) -> tuple[int, Any]:
        try:
            return 0, int(case_id)
        except ValueError:
            return 1, case_id

    return sorted(common, key=sort_key)


def original_value(data: WorkbookData, case_id: str, field: str) -> str:
    return scalar_to_text(data.original_records[case_id].get(field))


def collect_scalar_pairs(
    data_a: WorkbookData,
    data_b: WorkbookData,
    case_ids: Sequence[str],
    field: str,
) -> tuple[list[tuple[str, str]], list[str], list[str]]:
    paired: list[tuple[str, str]] = []
    missing_a: list[str] = []
    missing_b: list[str] = []
    for case_id in case_ids:
        key_a = canonical_key(field, data_a.records[case_id].get(field))
        key_b = canonical_key(field, data_b.records[case_id].get(field))
        if not key_a:
            missing_a.append(case_id)
        if not key_b:
            missing_b.append(case_id)
        if key_a and key_b:
            paired.append((key_a, key_b))
    return paired, missing_a, missing_b


def analyze_scalar_field(
    data_a: WorkbookData,
    data_b: WorkbookData,
    case_ids: Sequence[str],
    field: str,
    kind: str,
    order: Sequence[str] | None = None,
    bootstrap_iterations: int = 0,
    seed: int = 2026,
) -> dict[str, Any]:
    pairs, missing_a, missing_b = collect_scalar_pairs(data_a, data_b, case_ids, field)
    keys_a = [pair[0] for pair in pairs]
    keys_b = [pair[1] for pair in pairs]

    result: dict[str, Any] = {
        "field": field,
        "kind": kind,
        "n_aligned": len(case_ids),
        "n_paired": len(pairs),
        "missing_a": len(missing_a),
        "missing_b": len(missing_b),
        "exact": exact_agreement(keys_a, keys_b),
        "disagreements": sum(a != b for a, b in pairs),
        "kappa": cohen_kappa(keys_a, keys_b),
        "linear_weighted_kappa": None,
        "quadratic_weighted_kappa": None,
        "n_ordered": None,
        "excluded_from_weighted": None,
        "ci_lower": None,
        "ci_upper": None,
        "bootstrap_valid": 0,
        "confusion_labels": [],
        "confusion_matrix": [],
    }

    preferred_order = list(order or [])
    labels, matrix = confusion_matrix(keys_a, keys_b, preferred_order)
    result["confusion_labels"] = labels
    result["confusion_matrix"] = matrix

    if kind == "ordinal" and order:
        ordered_pairs = [pair for pair in pairs if pair[0] in order and pair[1] in order]
        ordered_a = [pair[0] for pair in ordered_pairs]
        ordered_b = [pair[1] for pair in ordered_pairs]
        result["n_ordered"] = len(ordered_pairs)
        result["excluded_from_weighted"] = len(pairs) - len(ordered_pairs)
        result["linear_weighted_kappa"] = weighted_kappa(
            ordered_a, ordered_b, order, weighting="linear"
        )
        result["quadratic_weighted_kappa"] = weighted_kappa(
            ordered_a, ordered_b, order, weighting="quadratic"
        )
        if bootstrap_iterations > 0:
            metric = lambda a, b: weighted_kappa(a, b, order, weighting="quadratic")
            lower, upper, valid = bootstrap_ci(
                ordered_pairs, metric, bootstrap_iterations, seed
            )
            result["ci_lower"] = lower
            result["ci_upper"] = upper
            result["bootstrap_valid"] = valid
    elif bootstrap_iterations > 0:
        lower, upper, valid = bootstrap_ci(
            pairs, cohen_kappa, bootstrap_iterations, seed
        )
        result["ci_lower"] = lower
        result["ci_upper"] = upper
        result["bootstrap_valid"] = valid

    return result


def analyze_multilabel_field(
    data_a: WorkbookData,
    data_b: WorkbookData,
    case_ids: Sequence[str],
    field: str,
) -> dict[str, Any]:
    sets_a: list[frozenset[str]] = []
    sets_b: list[frozenset[str]] = []
    missing_a = 0
    missing_b = 0
    for case_id in case_ids:
        set_a = parse_multilabel(data_a.records[case_id].get(field))
        set_b = parse_multilabel(data_b.records[case_id].get(field))
        if not set_a:
            missing_a += 1
        if not set_b:
            missing_b += 1
        sets_a.append(set_a)
        sets_b.append(set_b)
    metrics = multilabel_metrics(sets_a, sets_b)
    metrics.update(
        {
            "field": field,
            "missing_a": missing_a,
            "missing_b": missing_b,
            "disagreements": sum(a != b for a, b in zip(sets_a, sets_b)),
        }
    )
    return metrics


def analyze_text_field(
    data_a: WorkbookData,
    data_b: WorkbookData,
    case_ids: Sequence[str],
    field: str,
) -> dict[str, Any]:
    both_present = 0
    both_blank = 0
    one_blank = 0
    exact_count = 0
    for case_id in case_ids:
        value_a = normalized_text_key(data_a.records[case_id].get(field))
        value_b = normalized_text_key(data_b.records[case_id].get(field))
        if value_a and value_b:
            both_present += 1
            if value_a == value_b:
                exact_count += 1
        elif not value_a and not value_b:
            both_blank += 1
        else:
            one_blank += 1
    return {
        "field": field,
        "n_aligned": len(case_ids),
        "both_present": both_present,
        "both_blank": both_blank,
        "one_blank": one_blank,
        "normalized_exact_count": exact_count,
        "normalized_exact": safe_divide(exact_count, both_present),
    }


def get_disagreement_rows(
    data_a: WorkbookData,
    data_b: WorkbookData,
    case_ids: Sequence[str],
    scalar_fields: Sequence[str],
    multilabel_fields: Sequence[str],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for case_id in case_ids:
        for field in scalar_fields:
            key_a = canonical_key(field, data_a.records[case_id].get(field))
            key_b = canonical_key(field, data_b.records[case_id].get(field))
            if key_a != key_b:
                rows.append(
                    {
                        "Case ID": case_id,
                        "Field": field,
                        "Annotator A": original_value(data_a, case_id, field),
                        "Annotator B": original_value(data_b, case_id, field),
                        "Source Link(s) / Citation(s)": original_value(
                            data_a, case_id, "Source Link(s) / Citation(s)"
                        ),
                    }
                )
        for field in multilabel_fields:
            set_a = parse_multilabel(data_a.records[case_id].get(field))
            set_b = parse_multilabel(data_b.records[case_id].get(field))
            if set_a != set_b:
                rows.append(
                    {
                        "Case ID": case_id,
                        "Field": field,
                        "Annotator A": original_value(data_a, case_id, field),
                        "Annotator B": original_value(data_b, case_id, field),
                        "Source Link(s) / Citation(s)": original_value(
                            data_a, case_id, "Source Link(s) / Citation(s)"
                        ),
                    }
                )
    return rows


def distributions_for_field(
    data: WorkbookData,
    case_ids: Sequence[str],
    field: str,
    multilabel: bool = False,
) -> Counter[str]:
    counts: Counter[str] = Counter()
    for case_id in case_ids:
        if multilabel:
            values = parse_multilabel(data.records[case_id].get(field))
            if not values:
                counts["<blank>"] += 1
            else:
                counts.update(values)
        else:
            value = canonical_key(field, data.records[case_id].get(field))
            counts[value or "<blank>"] += 1
    return counts


def validate_vocabularies(
    data: WorkbookData,
    case_ids: Sequence[str],
    fields: Sequence[str],
    multilabel_fields: Sequence[str],
) -> list[dict[str, Any]]:
    issues: list[dict[str, Any]] = []
    for field in fields:
        allowed = data.vocabularies.get(field)
        if not allowed:
            continue
        allowed_keys = {canonical_key(field, value) for value in allowed}
        for case_id in case_ids:
            key = canonical_key(field, data.records[case_id].get(field))
            if key and key not in allowed_keys:
                issues.append(
                    {
                        "Workbook": data.path.name,
                        "Case ID": case_id,
                        "Field": field,
                        "Value": original_value(data, case_id, field),
                        "Issue": "Value not found in the workbook's Lists vocabulary",
                    }
                )
    for field in multilabel_fields:
        allowed = data.vocabularies.get(field)
        if not allowed:
            continue
        allowed_keys = {value.casefold() for value in allowed}
        for case_id in case_ids:
            for label in parse_multilabel(data.records[case_id].get(field)):
                if label not in allowed_keys:
                    issues.append(
                        {
                            "Workbook": data.path.name,
                            "Case ID": case_id,
                            "Field": field,
                            "Value": original_value(data, case_id, field),
                            "Issue": f"Label {label!r} not found in the workbook's Lists vocabulary",
                        }
                    )
    return issues


# ---------------------------------------------------------------------------
# Output helpers
# ---------------------------------------------------------------------------


def as_number(value: float | None) -> float | None:
    if value is None or not math.isfinite(value):
        return None
    return float(value)


def percentage(value: float | None) -> float | None:
    return None if value is None else value


def write_csv(path: Path, rows: Sequence[Mapping[str, Any]], headers: Sequence[str]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(headers), extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: row.get(header, "") for header in headers})


def json_safe(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]
    if isinstance(value, set):
        return sorted(json_safe(item) for item in value)
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    return value


def style_sheet(ws: Any, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells[: min(len(column_cells), 200)]:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        width = min(max(max_length + 2, 10), 42)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width


def write_table_sheet(
    workbook: Workbook,
    title: str,
    headers: Sequence[str],
    rows: Sequence[Sequence[Any]],
    percent_columns: Sequence[int] = (),
    decimal_columns: Sequence[int] = (),
) -> Any:
    ws = workbook.create_sheet(title=title[:31])
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    style_sheet(ws)
    for column in percent_columns:
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = "0.0%"
    for column in decimal_columns:
        for cell in ws.iter_cols(min_col=column, max_col=column, min_row=2):
            for item in cell:
                item.number_format = "0.000"
    return ws


def build_excel_report(
    output_path: Path,
    metadata: dict[str, Any],
    principal_results: Sequence[dict[str, Any]],
    supporting_results: Sequence[dict[str, Any]],
    multilabel_results: Sequence[dict[str, Any]],
    text_results: Sequence[dict[str, Any]],
    principal_disagreements: Sequence[dict[str, Any]],
    all_controlled_disagreements: Sequence[dict[str, Any]],
    distribution_rows: Sequence[dict[str, Any]],
    vocabulary_issues: Sequence[dict[str, Any]],
) -> None:
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Overview
    overview = workbook.create_sheet("Overview")
    overview.sheet_view.showGridLines = False
    overview["A1"] = "AIFTax Inter-Rater Agreement Report"
    overview["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    overview["A1"].fill = PatternFill("solid", fgColor="17365D")
    overview.merge_cells("A1:F1")
    overview["A3"] = "Annotator A"
    overview["B3"] = metadata["annotator_a"]
    overview["A4"] = "Annotator B"
    overview["B4"] = metadata["annotator_b"]
    overview["A5"] = "Aligned cases"
    overview["B5"] = metadata["aligned_cases"]
    overview["A6"] = "Generated"
    overview["B6"] = metadata["generated_utc"]
    overview["A9"] = "Principal analytical variables"
    overview["A9"].font = Font(bold=True, color="17365D")
    overview.append([])
    start_row = 10
    headers = [
        "Field",
        "Exact agreement",
        "Primary statistic",
        "Estimate",
        "95% CI lower",
        "95% CI upper",
    ]
    for col, header in enumerate(headers, 1):
        cell = overview.cell(start_row, col, header)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for offset, result in enumerate(principal_results, 1):
        estimate = (
            result["kappa"]
            if result["kind"] == "nominal"
            else result["quadratic_weighted_kappa"]
        )
        overview.append(
            [
                result["field"],
                result["exact"],
                PRINCIPAL_FIELDS[result["field"]]["primary_metric"],
                estimate,
                result["ci_lower"],
                result["ci_upper"],
            ]
        )
    for row in range(start_row + 1, start_row + 1 + len(principal_results)):
        overview.cell(row, 2).number_format = "0.0%"
        for col in (4, 5, 6):
            overview.cell(row, col).number_format = "0.000"
    overview["A17"] = "Interpretation note"
    overview["A17"].font = Font(bold=True, color="17365D")
    overview["A18"] = (
        "Kappa values are reported without automatic qualitative labels. "
        "Weighted kappas exclude IE, NA, Other, and other values outside the ordered scale. "
        "Narrative fields are reported only as normalized exact-match diagnostics."
    )
    overview.merge_cells("A18:F20")
    overview["A18"].alignment = Alignment(wrap_text=True, vertical="top")
    for column, width in {"A": 28, "B": 48, "C": 28, "D": 14, "E": 14, "F": 14}.items():
        overview.column_dimensions[column].width = width

    # Principal summary
    principal_rows = []
    for result in principal_results:
        primary_estimate = (
            result["kappa"]
            if result["kind"] == "nominal"
            else result["quadratic_weighted_kappa"]
        )
        principal_rows.append(
            [
                result["field"],
                result["kind"],
                result["n_aligned"],
                result["n_paired"],
                result["n_ordered"],
                result["exact"],
                result["kappa"],
                result["linear_weighted_kappa"],
                result["quadratic_weighted_kappa"],
                PRINCIPAL_FIELDS[result["field"]]["primary_metric"],
                primary_estimate,
                result["ci_lower"],
                result["ci_upper"],
                result["disagreements"],
                result["missing_a"],
                result["missing_b"],
                result["excluded_from_weighted"],
            ]
        )
    write_table_sheet(
        workbook,
        "Principal Agreement",
        [
            "Field",
            "Level",
            "N aligned",
            "N paired",
            "N ordered",
            "Exact agreement",
            "Unweighted kappa",
            "Linear weighted kappa",
            "Quadratic weighted kappa",
            "Primary statistic",
            "Primary estimate",
            "95% CI lower",
            "95% CI upper",
            "Disagreements",
            "Missing A",
            "Missing B",
            "Excluded from weighted",
        ],
        principal_rows,
        percent_columns=[6],
        decimal_columns=[7, 8, 9, 11, 12, 13],
    )

    # Supporting controlled fields
    supporting_rows = []
    for result in supporting_results:
        supporting_rows.append(
            [
                result["field"],
                result["kind"],
                result["n_paired"],
                result["exact"],
                result["kappa"],
                result["linear_weighted_kappa"],
                result["quadratic_weighted_kappa"],
                result["disagreements"],
                result["missing_a"],
                result["missing_b"],
                result["n_ordered"],
                result["excluded_from_weighted"],
            ]
        )
    write_table_sheet(
        workbook,
        "Supporting Controlled",
        [
            "Field",
            "Level",
            "N paired",
            "Exact agreement",
            "Unweighted kappa",
            "Linear weighted kappa",
            "Quadratic weighted kappa",
            "Disagreements",
            "Missing A",
            "Missing B",
            "N ordered",
            "Excluded from weighted",
        ],
        supporting_rows,
        percent_columns=[4],
        decimal_columns=[5, 6, 7],
    )

    # Multi-label summary
    multilabel_rows = []
    per_label_rows = []
    for result in multilabel_results:
        multilabel_rows.append(
            [
                result["field"],
                result["n"],
                result["exact"],
                result["mean_jaccard"],
                result["micro_f1"],
                result["macro_label_kappa"],
                result["label_count"],
                result["disagreements"],
                result["missing_a"],
                result["missing_b"],
            ]
        )
        for label_result in result["per_label"]:
            per_label_rows.append(
                [
                    result["field"],
                    label_result["label"],
                    label_result["a_positive"],
                    label_result["b_positive"],
                    label_result["binary_agreement"],
                    label_result["kappa"],
                ]
            )
    write_table_sheet(
        workbook,
        "Multilabel Agreement",
        [
            "Field",
            "N",
            "Exact set agreement",
            "Mean Jaccard",
            "Micro F1",
            "Macro per-label kappa",
            "Distinct labels",
            "Disagreements",
            "Missing A",
            "Missing B",
        ],
        multilabel_rows,
        percent_columns=[3, 4, 5],
        decimal_columns=[6],
    )
    write_table_sheet(
        workbook,
        "Per-label Kappa",
        ["Field", "Label", "A positive", "B positive", "Binary agreement", "Kappa"],
        per_label_rows,
        percent_columns=[5],
        decimal_columns=[6],
    )

    # Text diagnostics
    text_rows = [
        [
            result["field"],
            result["n_aligned"],
            result["both_present"],
            result["both_blank"],
            result["one_blank"],
            result["normalized_exact_count"],
            result["normalized_exact"],
            "Diagnostic only; not a categorical reliability statistic",
        ]
        for result in text_results
    ]
    write_table_sheet(
        workbook,
        "Text Diagnostics",
        [
            "Field",
            "N aligned",
            "Both present",
            "Both blank",
            "One blank",
            "Normalized exact count",
            "Normalized exact agreement",
            "Treatment",
        ],
        text_rows,
        percent_columns=[7],
    )

    # Disagreements
    disagreement_headers = [
        "Case ID",
        "Field",
        "Annotator A",
        "Annotator B",
        "Source Link(s) / Citation(s)",
    ]
    write_table_sheet(
        workbook,
        "Principal Disagreements",
        disagreement_headers,
        [[row.get(header) for header in disagreement_headers] for row in principal_disagreements],
    )
    write_table_sheet(
        workbook,
        "Controlled Disagreements",
        disagreement_headers,
        [
            [row.get(header) for header in disagreement_headers]
            for row in all_controlled_disagreements
        ],
    )

    # Distributions
    write_table_sheet(
        workbook,
        "Label Distributions",
        ["Field", "Label", "Annotator A count", "Annotator B count"],
        [
            [row["Field"], row["Label"], row["Annotator A count"], row["Annotator B count"]]
            for row in distribution_rows
        ],
    )

    # Validation
    validation_rows = [
        ["Annotator A duplicate Case IDs", ", ".join(metadata["duplicate_case_ids_a"]) or "None"],
        ["Annotator B duplicate Case IDs", ", ".join(metadata["duplicate_case_ids_b"]) or "None"],
        ["Cases only in Annotator A", ", ".join(metadata["cases_only_a"]) or "None"],
        ["Cases only in Annotator B", ", ".join(metadata["cases_only_b"]) or "None"],
        ["Source-link mismatches", metadata["source_link_mismatch_count"]],
        ["Vocabulary issues", len(vocabulary_issues)],
        ["Notes / Flag treatment", "Excluded from agreement calculations"],
    ]
    write_table_sheet(workbook, "Validation", ["Check", "Result"], validation_rows)

    if vocabulary_issues:
        issue_headers = ["Workbook", "Case ID", "Field", "Value", "Issue"]
        write_table_sheet(
            workbook,
            "Vocabulary Issues",
            issue_headers,
            [[row.get(header) for header in issue_headers] for row in vocabulary_issues],
        )

    # Confusion matrices for principal variables
    for result in principal_results:
        title_map = {
            "Failure Category": "CM Failure Category",
            "Propagation Reach": "CM Propagation Reach",
            "Risk Assessment": "CM Risk Assessment",
            "Recovery Complexity": "CM Recovery Complexity",
        }
        ws = workbook.create_sheet(title_map[result["field"]])
        labels = result["confusion_labels"]
        ws.cell(1, 1, "Annotator A \\ Annotator B")
        for index, label in enumerate(labels, 2):
            ws.cell(1, index, label)
            ws.cell(index, 1, label)
        for row_index, row in enumerate(result["confusion_matrix"], 2):
            for col_index, value in enumerate(row, 2):
                ws.cell(row_index, col_index, value)
        style_sheet(ws, freeze="B2")

    workbook.save(output_path)


# ---------------------------------------------------------------------------
# Main program
# ---------------------------------------------------------------------------


def parse_args(argv: Sequence[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Calculate AIFTax inter-rater annotation agreement."
    )
    parser.add_argument("--annotator-a", required=True, type=Path)
    parser.add_argument("--annotator-b", required=True, type=Path)
    parser.add_argument("--sheet", default="Annotations")
    parser.add_argument("--output-dir", type=Path, default=Path("irr_results"))
    parser.add_argument("--output-prefix", default="AIFTax_IRR")
    parser.add_argument(
        "--bootstrap",
        type=int,
        default=2000,
        help="Bootstrap iterations for 95%% CIs on principal kappa estimates (default: 2000).",
    )
    parser.add_argument("--seed", type=int, default=2026)
    return parser.parse_args(argv)


def main(argv: Sequence[str] | None = None) -> int:
    args = parse_args(argv)
    if not args.annotator_a.exists():
        raise SystemExit(f"Annotator A file not found: {args.annotator_a}")
    if not args.annotator_b.exists():
        raise SystemExit(f"Annotator B file not found: {args.annotator_b}")
    if args.bootstrap < 0:
        raise SystemExit("--bootstrap must be zero or greater")

    args.output_dir.mkdir(parents=True, exist_ok=True)

    data_a = read_annotation_workbook(args.annotator_a, args.sheet)
    data_b = read_annotation_workbook(args.annotator_b, args.sheet)

    case_ids = paired_case_ids(data_a, data_b)
    cases_only_a = sorted(set(data_a.records) - set(data_b.records))
    cases_only_b = sorted(set(data_b.records) - set(data_a.records))
    if not case_ids:
        raise SystemExit("No common Case IDs were found between the two workbooks")

    missing_required_headers = []
    required_fields = (
        ADMINISTRATIVE_FIELDS
        + list(PRINCIPAL_FIELDS)
        + SUPPORTING_NOMINAL_FIELDS
        + list(SUPPORTING_ORDINAL_FIELDS)
        + MULTILABEL_FIELDS
        + TEXT_DIAGNOSTIC_FIELDS
    )
    for field in required_fields:
        if field not in data_a.headers:
            missing_required_headers.append(f"Annotator A: {field}")
        if field not in data_b.headers:
            missing_required_headers.append(f"Annotator B: {field}")
    if missing_required_headers:
        raise SystemExit(
            "Required annotation columns are missing:\n  - "
            + "\n  - ".join(missing_required_headers)
        )

    source_link_mismatches = []
    for case_id in case_ids:
        link_a = normalized_text_key(
            data_a.records[case_id].get("Source Link(s) / Citation(s)")
        )
        link_b = normalized_text_key(
            data_b.records[case_id].get("Source Link(s) / Citation(s)")
        )
        if link_a != link_b:
            source_link_mismatches.append(case_id)

    principal_results = []
    for field_index, (field, config) in enumerate(PRINCIPAL_FIELDS.items()):
        result = analyze_scalar_field(
            data_a,
            data_b,
            case_ids,
            field,
            config["kind"],
            config.get("order"),
            args.bootstrap,
            args.seed + field_index,
        )
        principal_results.append(result)

    supporting_results = []
    nominal_fields = list(SUPPORTING_NOMINAL_FIELDS)
    for field in nominal_fields:
        supporting_results.append(
            analyze_scalar_field(data_a, data_b, case_ids, field, "nominal")
        )
    for field, order in SUPPORTING_ORDINAL_FIELDS.items():
        supporting_results.append(
            analyze_scalar_field(data_a, data_b, case_ids, field, "ordinal", order)
        )

    multilabel_results = [
        analyze_multilabel_field(data_a, data_b, case_ids, field)
        for field in MULTILABEL_FIELDS
    ]
    text_results = [
        analyze_text_field(data_a, data_b, case_ids, field)
        for field in TEXT_DIAGNOSTIC_FIELDS
    ]

    principal_disagreements = get_disagreement_rows(
        data_a,
        data_b,
        case_ids,
        list(PRINCIPAL_FIELDS),
        [],
    )
    all_controlled_disagreements = get_disagreement_rows(
        data_a,
        data_b,
        case_ids,
        list(PRINCIPAL_FIELDS) + nominal_fields + list(SUPPORTING_ORDINAL_FIELDS),
        MULTILABEL_FIELDS,
    )

    distribution_rows: list[dict[str, Any]] = []
    scalar_distribution_fields = (
        list(PRINCIPAL_FIELDS) + nominal_fields + list(SUPPORTING_ORDINAL_FIELDS)
    )
    for field in scalar_distribution_fields:
        counts_a = distributions_for_field(data_a, case_ids, field)
        counts_b = distributions_for_field(data_b, case_ids, field)
        for label in sorted(set(counts_a) | set(counts_b)):
            distribution_rows.append(
                {
                    "Field": field,
                    "Label": label,
                    "Annotator A count": counts_a[label],
                    "Annotator B count": counts_b[label],
                }
            )
    for field in MULTILABEL_FIELDS:
        counts_a = distributions_for_field(data_a, case_ids, field, multilabel=True)
        counts_b = distributions_for_field(data_b, case_ids, field, multilabel=True)
        for label in sorted(set(counts_a) | set(counts_b)):
            distribution_rows.append(
                {
                    "Field": field,
                    "Label": label,
                    "Annotator A count": counts_a[label],
                    "Annotator B count": counts_b[label],
                }
            )

    vocabulary_fields = nominal_fields + list(PRINCIPAL_FIELDS) + list(
        SUPPORTING_ORDINAL_FIELDS
    )
    vocabulary_issues = validate_vocabularies(
        data_a, case_ids, vocabulary_fields, MULTILABEL_FIELDS
    ) + validate_vocabularies(data_b, case_ids, vocabulary_fields, MULTILABEL_FIELDS)

    metadata = {
        "annotator_a": str(args.annotator_a.resolve()),
        "annotator_b": str(args.annotator_b.resolve()),
        "sheet": args.sheet,
        "header_row_a": data_a.header_row,
        "header_row_b": data_b.header_row,
        "aligned_cases": len(case_ids),
        "case_ids": case_ids,
        "cases_only_a": cases_only_a,
        "cases_only_b": cases_only_b,
        "duplicate_case_ids_a": data_a.duplicate_case_ids,
        "duplicate_case_ids_b": data_b.duplicate_case_ids,
        "source_link_mismatch_count": len(source_link_mismatches),
        "source_link_mismatch_case_ids": source_link_mismatches,
        "bootstrap_iterations": args.bootstrap,
        "bootstrap_seed": args.seed,
        "generated_utc": datetime.now(timezone.utc).isoformat(timespec="seconds"),
    }

    report = {
        "metadata": metadata,
        "principal": principal_results,
        "supporting_controlled": supporting_results,
        "multilabel": multilabel_results,
        "text_diagnostics": text_results,
        "vocabulary_issues": vocabulary_issues,
    }

    prefix = args.output_prefix
    excel_path = args.output_dir / f"{prefix}_Report.xlsx"
    json_path = args.output_dir / f"{prefix}_Report.json"
    principal_csv_path = args.output_dir / f"{prefix}_Principal_Summary.csv"
    supporting_csv_path = args.output_dir / f"{prefix}_Supporting_Summary.csv"
    multilabel_csv_path = args.output_dir / f"{prefix}_Multilabel_Summary.csv"
    principal_disagreement_path = args.output_dir / f"{prefix}_Principal_Disagreements.csv"
    controlled_disagreement_path = args.output_dir / f"{prefix}_Controlled_Disagreements.csv"

    build_excel_report(
        excel_path,
        metadata,
        principal_results,
        supporting_results,
        multilabel_results,
        text_results,
        principal_disagreements,
        all_controlled_disagreements,
        distribution_rows,
        vocabulary_issues,
    )

    with json_path.open("w", encoding="utf-8") as handle:
        json.dump(json_safe(report), handle, indent=2, ensure_ascii=False)

    principal_csv_rows = []
    for result in principal_results:
        primary_estimate = (
            result["kappa"]
            if result["kind"] == "nominal"
            else result["quadratic_weighted_kappa"]
        )
        principal_csv_rows.append(
            {
                "Field": result["field"],
                "Level": result["kind"],
                "N paired": result["n_paired"],
                "N ordered": result["n_ordered"],
                "Exact agreement": result["exact"],
                "Unweighted kappa": result["kappa"],
                "Linear weighted kappa": result["linear_weighted_kappa"],
                "Quadratic weighted kappa": result["quadratic_weighted_kappa"],
                "Primary statistic": PRINCIPAL_FIELDS[result["field"]]["primary_metric"],
                "Primary estimate": primary_estimate,
                "95% CI lower": result["ci_lower"],
                "95% CI upper": result["ci_upper"],
                "Disagreements": result["disagreements"],
            }
        )
    write_csv(
        principal_csv_path,
        principal_csv_rows,
        [
            "Field",
            "Level",
            "N paired",
            "N ordered",
            "Exact agreement",
            "Unweighted kappa",
            "Linear weighted kappa",
            "Quadratic weighted kappa",
            "Primary statistic",
            "Primary estimate",
            "95% CI lower",
            "95% CI upper",
            "Disagreements",
        ],
    )

    supporting_csv_rows = [
        {
            "Field": result["field"],
            "Level": result["kind"],
            "N paired": result["n_paired"],
            "Exact agreement": result["exact"],
            "Unweighted kappa": result["kappa"],
            "Linear weighted kappa": result["linear_weighted_kappa"],
            "Quadratic weighted kappa": result["quadratic_weighted_kappa"],
            "Disagreements": result["disagreements"],
        }
        for result in supporting_results
    ]
    write_csv(
        supporting_csv_path,
        supporting_csv_rows,
        [
            "Field",
            "Level",
            "N paired",
            "Exact agreement",
            "Unweighted kappa",
            "Linear weighted kappa",
            "Quadratic weighted kappa",
            "Disagreements",
        ],
    )

    multilabel_csv_rows = [
        {
            "Field": result["field"],
            "N": result["n"],
            "Exact set agreement": result["exact"],
            "Mean Jaccard": result["mean_jaccard"],
            "Micro F1": result["micro_f1"],
            "Macro per-label kappa": result["macro_label_kappa"],
            "Distinct labels": result["label_count"],
            "Disagreements": result["disagreements"],
        }
        for result in multilabel_results
    ]
    write_csv(
        multilabel_csv_path,
        multilabel_csv_rows,
        [
            "Field",
            "N",
            "Exact set agreement",
            "Mean Jaccard",
            "Micro F1",
            "Macro per-label kappa",
            "Distinct labels",
            "Disagreements",
        ],
    )

    disagreement_headers = [
        "Case ID",
        "Field",
        "Annotator A",
        "Annotator B",
        "Source Link(s) / Citation(s)",
    ]
    write_csv(principal_disagreement_path, principal_disagreements, disagreement_headers)
    write_csv(controlled_disagreement_path, all_controlled_disagreements, disagreement_headers)

    print(f"Aligned cases: {len(case_ids)}")
    print(f"Source-link mismatches: {len(source_link_mismatches)}")
    print("Principal agreement:")
    for result in principal_results:
        estimate = (
            result["kappa"]
            if result["kind"] == "nominal"
            else result["quadratic_weighted_kappa"]
        )
        exact = "NA" if result["exact"] is None else f"{result['exact']:.3f}"
        metric = "NA" if estimate is None else f"{estimate:.3f}"
        print(f"  {result['field']}: exact={exact}, primary kappa={metric}")
    print(f"Excel report: {excel_path}")
    print(f"JSON report: {json_path}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        raise SystemExit("Interrupted")
